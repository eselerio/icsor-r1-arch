"""Workbook and runtime contract tests for the ASM2D-TSN reference model."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
from openpyxl import load_workbook

from src.models.simulation.asm2d_tsn_simulation import (
    _build_influent_state_sample,
    _generate_lhs_candidate_pool,
    create_asm2d_tsn_workbook,
    generate_asm2d_tsn_dataset,
    get_asm2d_tsn_matrices,
    load_asm2d_tsn_workbook_composition,
    load_asm2d_tsn_simulation_params,
    resolve_asm2d_tsn_simulation_artifact_paths,
    resolve_asm2d_tsn_workbook_path,
    run_asm2d_tsn_simulation,
    simulate_asm2d_tsn_steady_state,
    sweep_asm2d_tsn_operating_space,
)


def _column_index_by_header(worksheet) -> dict[str, int]:
    return {
        str(cell.value): index
        for index, cell in enumerate(worksheet[1], start=1)
        if cell.value is not None
    }


def _row_index_by_value(worksheet, column_number: int) -> dict[str, int]:
    index: dict[str, int] = {}
    for row_number in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=column_number).value
        if value is not None:
            index[str(value)] = row_number
    return index


def _build_midpoint_influent_state(model_params: dict[str, object]) -> np.ndarray:
    state_columns = list(model_params["workbook"]["state_columns"])
    midpoint_sample = np.array(
        [
            np.mean(model_params["influent_state_ranges"][state_name])
            for state_name in state_columns
        ],
        dtype=float,
    )
    return _build_influent_state_sample(midpoint_sample)


def _build_temp_cache_paths_config(tmp_dir: Path) -> dict[str, str]:
    cache_root = (tmp_dir / "cache").as_posix()
    return {
        "asm2d_tsn_composition_cache_pattern": f"{cache_root}/composition_matrix_cache_{{workbook_hash}}.pkl",
        "asm2d_tsn_composition_cache_metadata_pattern": f"{cache_root}/composition_matrix_cache_{{workbook_hash}}.json",
    }


class Asm2dTsnWorkbookTests(unittest.TestCase):
    def test_resolve_workbook_path_uses_configured_location(self) -> None:
        workbook_path = resolve_asm2d_tsn_workbook_path()

        self.assertTrue(workbook_path.as_posix().endswith("data/asm2d-tsn/asm2d_tsn_workbook.xlsx"))

    def test_workbook_config_contains_expected_dimensions(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        workbook_config = params["workbook"]
        workbook_composition = load_asm2d_tsn_workbook_composition(
            model_params=params,
            use_cache=False,
        )

        self.assertEqual(
            workbook_config["sheets"],
            ["stoichiometric_matrix", "composition_matrix", "parameter_table"],
        )
        self.assertEqual(len(workbook_config["processes"]), 28)
        self.assertEqual(len(workbook_config["state_columns"]), 20)
        self.assertEqual(workbook_composition["state_columns"], workbook_config["state_columns"])
        self.assertGreaterEqual(len(workbook_composition["measured_output_columns"]), 1)
        self.assertNotIn("X_TSS", workbook_config["state_columns"])

    def test_create_workbook_writes_required_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = create_asm2d_tsn_workbook(Path(tmp_dir) / "asm2d_tsn.xlsx")
            workbook = load_workbook(workbook_path, data_only=False)

        self.assertEqual(
            workbook.sheetnames,
            ["stoichiometric_matrix", "composition_matrix", "parameter_table"],
        )

    def test_stoichiometric_matrix_contains_formula_links(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = create_asm2d_tsn_workbook(Path(tmp_dir) / "asm2d_tsn.xlsx")
            workbook = load_workbook(workbook_path, data_only=False)

        worksheet = workbook["stoichiometric_matrix"]
        header_index = _column_index_by_header(worksheet)
        process_row_index = _row_index_by_value(worksheet, 2)

        aerobic_hydrolysis_row = process_row_index["Aerobic hydrolysis"]
        precipitation_row = process_row_index["Precipitation"]
        aob_growth_row = process_row_index["Aerobic growth of X_AOB"]

        self.assertIn("parameter_table", str(worksheet.cell(aerobic_hydrolysis_row, header_index["S_F"]).value))
        self.assertIn("parameter_table", str(worksheet.cell(aerobic_hydrolysis_row, header_index["S_NH4"]).value))
        self.assertIn("parameter_table", str(worksheet.cell(aob_growth_row, header_index["S_NO2"]).value))
        self.assertEqual(worksheet.cell(precipitation_row, header_index["S_PO4"]).value, "=-1")
        self.assertEqual(worksheet.cell(precipitation_row, header_index["X_MeOH"]).value, "=-3.45")
        self.assertEqual(worksheet.cell(precipitation_row, header_index["X_MeP"]).value, "=4.87")

    def test_composition_matrix_contains_formula_links(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = create_asm2d_tsn_workbook(Path(tmp_dir) / "asm2d_tsn.xlsx")
            workbook = load_workbook(workbook_path, data_only=False)

        worksheet = workbook["composition_matrix"]
        header_index = _column_index_by_header(worksheet)
        state_row_index = _row_index_by_value(worksheet, 2)

        self.assertIn("parameter_table", str(worksheet.cell(state_row_index["X_H"], header_index["TN"]).value))
        self.assertIn("parameter_table", str(worksheet.cell(state_row_index["X_MeP"], header_index["TP"]).value))
        self.assertIn("parameter_table", str(worksheet.cell(state_row_index["X_H"], header_index["TSS"]).value))
        self.assertEqual(worksheet.cell(state_row_index["X_MeOH"], header_index["TSS"]).value, "=1")
        self.assertEqual(worksheet.cell(state_row_index["X_MeP"], header_index["TSS"]).value, "=1")

    def test_workbook_composition_loader_tracks_added_composite_columns(self) -> None:
        params = load_asm2d_tsn_simulation_params()

        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_root = Path(tmp_dir)
            workbook_path = create_asm2d_tsn_workbook(temp_root / "asm2d_tsn.xlsx")
            workbook = load_workbook(workbook_path)
            worksheet = workbook["composition_matrix"]
            extra_column = worksheet.max_column + 1
            worksheet.cell(row=1, column=extra_column, value="TOC")
            for row_number in range(2, worksheet.max_row + 1):
                state_name = worksheet.cell(row=row_number, column=2).value
                if state_name is None:
                    continue
                worksheet.cell(row=row_number, column=extra_column, value="=0")
            workbook.save(workbook_path)
            workbook.close()

            composition_bundle = load_asm2d_tsn_workbook_composition(
                workbook_path=workbook_path,
                model_params=params,
                paths_config=_build_temp_cache_paths_config(temp_root),
                use_cache=False,
            )

        self.assertIn("TOC", composition_bundle["measured_output_columns"])
        self.assertEqual(
            composition_bundle["composition_matrix"].shape,
            (len(composition_bundle["measured_output_columns"]), len(composition_bundle["state_columns"])),
        )

    def test_workbook_composition_loader_tracks_removed_composite_columns(self) -> None:
        params = load_asm2d_tsn_simulation_params()

        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_root = Path(tmp_dir)
            workbook_path = create_asm2d_tsn_workbook(temp_root / "asm2d_tsn.xlsx")
            workbook = load_workbook(workbook_path)
            worksheet = workbook["composition_matrix"]
            header_index = _column_index_by_header(worksheet)
            worksheet.delete_cols(header_index["TKN"])
            workbook.save(workbook_path)
            workbook.close()

            composition_bundle = load_asm2d_tsn_workbook_composition(
                workbook_path=workbook_path,
                model_params=params,
                paths_config=_build_temp_cache_paths_config(temp_root),
                use_cache=False,
            )

        self.assertNotIn("TKN", composition_bundle["measured_output_columns"])

    def test_workbook_composition_cache_invalidates_after_workbook_change(self) -> None:
        params = load_asm2d_tsn_simulation_params()

        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_root = Path(tmp_dir)
            workbook_path = create_asm2d_tsn_workbook(temp_root / "asm2d_tsn.xlsx")
            cache_config = _build_temp_cache_paths_config(temp_root)

            first_bundle = load_asm2d_tsn_workbook_composition(
                workbook_path=workbook_path,
                model_params=params,
                paths_config=cache_config,
                use_cache=True,
            )
            second_bundle = load_asm2d_tsn_workbook_composition(
                workbook_path=workbook_path,
                model_params=params,
                paths_config=cache_config,
                use_cache=True,
            )

            workbook = load_workbook(workbook_path)
            worksheet = workbook["composition_matrix"]
            header_index = _column_index_by_header(worksheet)
            state_row_index = _row_index_by_value(worksheet, 2)
            worksheet.cell(row=state_row_index["X_MeP"], column=header_index["TP"], value="=2")
            workbook.save(workbook_path)
            workbook.close()

            third_bundle = load_asm2d_tsn_workbook_composition(
                workbook_path=workbook_path,
                model_params=params,
                paths_config=cache_config,
                use_cache=True,
            )

        self.assertEqual(first_bundle["cache_source"], "workbook")
        self.assertEqual(second_bundle["cache_source"], "cache")
        self.assertEqual(third_bundle["cache_source"], "workbook")
        self.assertNotEqual(first_bundle["workbook_sha256"], third_bundle["workbook_sha256"])


class Asm2dTsnSimulationTests(unittest.TestCase):
    def test_resolve_simulation_artifact_paths_use_requested_folder(self) -> None:
        dataset_path, metadata_path, dataset_relative = resolve_asm2d_tsn_simulation_artifact_paths(
            timestamp="20260330_000000"
        )

        self.assertTrue(dataset_path.as_posix().endswith("data/asm2d-tsn/simulation/data_20260330_000000.csv"))
        self.assertTrue(metadata_path.as_posix().endswith("data/asm2d-tsn/simulation/metadata_20260330_000000.json"))
        self.assertEqual(dataset_relative, "data/asm2d-tsn/simulation/data_20260330_000000.csv")

    def test_numeric_matrices_have_expected_shapes(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        matrix_bundle = get_asm2d_tsn_matrices(params)
        process_count = len(params["workbook"]["processes"])
        state_count = len(params["workbook"]["state_columns"])
        measured_output_count = len(matrix_bundle["measured_output_columns"])

        self.assertEqual(matrix_bundle["petersen_matrix"].shape, (process_count, state_count))
        self.assertEqual(matrix_bundle["composition_matrix"].shape, (measured_output_count, state_count))
        self.assertGreaterEqual(measured_output_count, 1)

    def test_generate_dataset_reports_fraction_and_composite_columns(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        dataset, metadata, matrix_bundle = generate_asm2d_tsn_dataset(
            model_params=params,
            n_samples=12,
            random_seed=7,
            parallel_workers=1,
        )
        state_columns = list(params["workbook"]["state_columns"])
        measured_output_columns = list(metadata["measured_output_columns"])
        expected_influent_fraction = [f"In_{name}" for name in state_columns]
        expected_influent_composite = [f"In_{name}" for name in measured_output_columns]
        expected_effluent_fraction = [f"Out_{name}" for name in state_columns]
        expected_independent = metadata["independent_columns"]
        expected_dependent = metadata["dependent_columns"]
        expected_ignored = metadata["ignored_columns"]

        self.assertEqual(expected_independent, metadata["operational_columns"] + expected_influent_fraction)
        self.assertEqual(expected_dependent, [f"Out_{name}" for name in measured_output_columns])
        self.assertEqual(expected_ignored, expected_influent_composite + expected_effluent_fraction)
        self.assertEqual(
            dataset.shape,
            (12, len(expected_independent) + len(expected_ignored) + len(expected_dependent)),
        )
        self.assertEqual(list(dataset.columns), expected_independent + expected_ignored + expected_dependent)
        self.assertTrue(all(column_name in dataset.columns for column_name in expected_influent_composite))
        self.assertTrue(all(column_name in dataset.columns for column_name in expected_effluent_fraction))
        self.assertFalse(any(column_name.startswith("Out_S_") or column_name.startswith("Out_X_") for column_name in expected_dependent))
        self.assertEqual(matrix_bundle["petersen_matrix"].shape, (len(params["workbook"]["processes"]), len(state_columns)))
        self.assertEqual(matrix_bundle["composition_matrix"].shape, (len(measured_output_columns), len(state_columns)))
        self.assertEqual(
            metadata["composition_source"]["workbook_sha256"],
            matrix_bundle["composition_workbook_sha256"],
        )

    def test_single_operating_point_solves_to_small_residual(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        influent_state = _build_midpoint_influent_state(params)

        solution, diagnostics = simulate_asm2d_tsn_steady_state(
            influent_state=influent_state,
            hrt_hours=24.0,
            aeration=1.5,
            model_params=params,
        )

        self.assertTrue(diagnostics["success"])
        self.assertTrue(diagnostics["accepted"])
        self.assertLessEqual(diagnostics["residual_max"], diagnostics["acceptance_threshold"])
        self.assertTrue((solution >= 0.0).all())

    def test_steady_state_responds_to_aeration_and_hrt(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        matrix_bundle = get_asm2d_tsn_matrices(params)
        influent_state = _build_midpoint_influent_state(params)
        state_index = dict(matrix_bundle["state_index"])
        output_index = {
            name: position for position, name in enumerate(matrix_bundle["measured_output_columns"])
        }

        low_aeration_state, _ = simulate_asm2d_tsn_steady_state(
            influent_state=influent_state,
            hrt_hours=24.0,
            aeration=0.75,
            model_params=params,
            matrix_bundle=matrix_bundle,
        )
        high_aeration_state, _ = simulate_asm2d_tsn_steady_state(
            influent_state=influent_state,
            hrt_hours=24.0,
            aeration=2.25,
            model_params=params,
            matrix_bundle=matrix_bundle,
        )

        self.assertGreater(high_aeration_state[state_index["S_O"]], low_aeration_state[state_index["S_O"]])
        self.assertGreater(high_aeration_state[state_index["S_NO3"]], low_aeration_state[state_index["S_NO3"]])

        low_hrt_state, _ = simulate_asm2d_tsn_steady_state(
            influent_state=influent_state,
            hrt_hours=12.0,
            aeration=1.5,
            model_params=params,
            matrix_bundle=matrix_bundle,
        )
        high_hrt_state, _ = simulate_asm2d_tsn_steady_state(
            influent_state=influent_state,
            hrt_hours=36.0,
            aeration=1.5,
            model_params=params,
            matrix_bundle=matrix_bundle,
        )
        low_hrt_outputs = matrix_bundle["composition_matrix"] @ low_hrt_state
        high_hrt_outputs = matrix_bundle["composition_matrix"] @ high_hrt_state

        self.assertLess(high_hrt_outputs[output_index["COD"]], low_hrt_outputs[output_index["COD"]])
        self.assertLess(high_hrt_state[state_index["S_NH4"]], low_hrt_state[state_index["S_NH4"]])

    def test_generate_dataset_retries_solver_failures(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        original_solver = simulate_asm2d_tsn_steady_state
        fail_once = {"remaining": 1}

        def flaky_solver(*args, **kwargs):
            if fail_once["remaining"] > 0:
                fail_once["remaining"] -= 1
                raise RuntimeError("transient steady-state failure")
            return original_solver(*args, **kwargs)

        with patch(
            "src.models.simulation.asm2d_tsn_simulation.simulate_asm2d_tsn_steady_state",
            side_effect=flaky_solver,
        ):
            dataset, metadata, _ = generate_asm2d_tsn_dataset(
                model_params=params,
                n_samples=1,
                random_seed=7,
                parallel_workers=1,
            )

        self.assertEqual(dataset.shape[0], 1)
        self.assertEqual(metadata["n_samples"], 1)
        self.assertEqual(fail_once["remaining"], 0)

    def test_run_simulation_returns_notebook_facing_bundle(self) -> None:
        result = run_asm2d_tsn_simulation(
            save_artifacts=False,
            n_samples=8,
            random_seed=5,
            parallel_workers=1,
        )

        self.assertIn("dataset", result)
        self.assertIn("metadata", result)
        self.assertIn("petersen_matrix", result)
        self.assertIn("composition_matrix", result)
        self.assertIn("matrix_bundle", result)
        self.assertIn("artifact_paths", result)
        self.assertEqual(result["artifact_paths"]["dataset_csv"], None)
        self.assertEqual(result["artifact_paths"]["metadata_json"], None)
        self.assertEqual(
            result["metadata"]["measured_output_columns"],
            result["matrix_bundle"]["measured_output_columns"],
        )

    def test_run_simulation_can_return_in_memory_debug_payloads(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        result = run_asm2d_tsn_simulation(
            save_artifacts=False,
            n_samples=4,
            random_seed=5,
            parallel_workers=1,
            include_debug_data=True,
            show_progress=False,
        )

        self.assertEqual(result["dataset"].shape[0], 4)
        self.assertIsNotNone(result["effluent_states"])
        self.assertIsNotNone(result["solver_diagnostics"])
        self.assertIsNotNone(result["solver_summary"])
        self.assertEqual(result["effluent_states"].shape, (4, len(params["workbook"]["state_columns"])))
        self.assertEqual(len(result["solver_diagnostics"]), 4)
        self.assertEqual(result["solver_summary"]["sample_count"], 4)
        self.assertIn("selected_strategy", result["solver_diagnostics"].columns)
        self.assertIn("dynamic_relaxation_used", result["solver_diagnostics"].columns)

    def test_operating_space_sweep_returns_calibration_summary(self) -> None:
        sweep_result = sweep_asm2d_tsn_operating_space(
            n_samples=16,
            random_seed=7,
            show_progress=False,
        )

        self.assertEqual(sweep_result["influent_states"].shape[0], 16)
        self.assertEqual(sweep_result["operating_conditions"].shape[0], 16)
        self.assertEqual(sweep_result["effluent_states"].shape[0], 16)
        self.assertEqual(len(sweep_result["solver_diagnostics"]), 16)
        self.assertEqual(sweep_result["summary"]["sample_count"], 16)
        self.assertGreaterEqual(sweep_result["summary"]["accepted_count"], 1)
        self.assertIn("residual_max_quantiles", sweep_result["summary"])
        self.assertIn("selected_strategy_counts", sweep_result["summary"])


class Asm2dTsnLhsSamplingTests(unittest.TestCase):
    def test_generate_lhs_candidate_pool_returns_correct_shape(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        state_columns = list(params["workbook"]["state_columns"])
        n_points = 20
        pool = _generate_lhs_candidate_pool(
            seed=0, n_points=n_points, ordered_names=state_columns, ranges=params["influent_state_ranges"]
        )

        self.assertEqual(pool.shape, (n_points, len(state_columns)))

    def test_generate_lhs_candidate_pool_is_within_configured_bounds(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        state_columns = list(params["workbook"]["state_columns"])
        pool = _generate_lhs_candidate_pool(
            seed=42, n_points=50, ordered_names=state_columns, ranges=params["influent_state_ranges"]
        )

        for col_idx, name in enumerate(state_columns):
            lo, hi = params["influent_state_ranges"][name]
            self.assertTrue(
                np.all(pool[:, col_idx] >= float(lo) - 1e-9),
                msg=f"{name} samples below lower bound",
            )
            self.assertTrue(
                np.all(pool[:, col_idx] <= float(hi) + 1e-9),
                msg=f"{name} samples above upper bound",
            )

    def test_generate_lhs_candidate_pool_is_deterministic_with_same_seed(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        state_columns = list(params["workbook"]["state_columns"])
        pool_a = _generate_lhs_candidate_pool(
            seed=7, n_points=10, ordered_names=state_columns, ranges=params["influent_state_ranges"]
        )
        pool_b = _generate_lhs_candidate_pool(
            seed=7, n_points=10, ordered_names=state_columns, ranges=params["influent_state_ranges"]
        )

        np.testing.assert_array_equal(pool_a, pool_b)

    def test_generate_lhs_candidate_pool_differs_across_seeds(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        state_columns = list(params["workbook"]["state_columns"])
        pool_a = _generate_lhs_candidate_pool(
            seed=1, n_points=10, ordered_names=state_columns, ranges=params["influent_state_ranges"]
        )
        pool_b = _generate_lhs_candidate_pool(
            seed=2, n_points=10, ordered_names=state_columns, ranges=params["influent_state_ranges"]
        )

        self.assertFalse(np.array_equal(pool_a, pool_b))

    def test_generate_lhs_candidate_pool_handles_degenerate_bounds(self) -> None:
        degenerate_ranges = {"x": [3.5, 3.5]}
        pool = _generate_lhs_candidate_pool(
            seed=0, n_points=5, ordered_names=["x"], ranges=degenerate_ranges
        )

        self.assertEqual(pool.shape, (5, 1))
        np.testing.assert_allclose(pool[:, 0], 3.5)

    def test_generate_lhs_candidate_pool_returns_empty_for_zero_points(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        state_columns = list(params["workbook"]["state_columns"])
        pool = _generate_lhs_candidate_pool(
            seed=0, n_points=0, ordered_names=state_columns, ranges=params["influent_state_ranges"]
        )

        self.assertEqual(pool.shape, (0, len(state_columns)))

    def test_generate_dataset_metadata_reports_latin_hypercube_sampling(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        _, metadata, _ = generate_asm2d_tsn_dataset(
            model_params=params, n_samples=4, random_seed=7, parallel_workers=1
        )

        self.assertEqual(metadata["sampling_method"], "latin_hypercube")

    def test_generate_dataset_is_reproducible_with_same_seed(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        dataset_a, _, _ = generate_asm2d_tsn_dataset(
            model_params=params, n_samples=8, random_seed=11, parallel_workers=1
        )
        dataset_b, _, _ = generate_asm2d_tsn_dataset(
            model_params=params, n_samples=8, random_seed=11, parallel_workers=1
        )

        pd = __import__("pandas")
        pd.testing.assert_frame_equal(dataset_a, dataset_b)

    def test_generate_dataset_samples_within_configured_ranges(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        dataset, metadata, _ = generate_asm2d_tsn_dataset(
            model_params=params, n_samples=10, random_seed=13, parallel_workers=1
        )
        influent_fraction_columns = metadata["influent_fraction_columns"]
        state_columns = list(params["workbook"]["state_columns"])
        operational_columns = list(params["operational_columns"])

        for state_name, col_name in zip(state_columns, influent_fraction_columns):
            lo, hi = params["influent_state_ranges"][state_name]
            col = dataset[col_name].to_numpy()
            self.assertTrue(np.all(col >= float(lo) - 1e-9), msg=f"{col_name} below lower bound")
            self.assertTrue(np.all(col <= float(hi) + 1e-9), msg=f"{col_name} above upper bound")

        for op_name in operational_columns:
            lo, hi = params["operational_ranges"][op_name]
            col = dataset[op_name].to_numpy()
            self.assertTrue(np.all(col >= float(lo) - 1e-9), msg=f"{op_name} below lower bound")
            self.assertTrue(np.all(col <= float(hi) + 1e-9), msg=f"{op_name} above upper bound")

    def test_sweep_is_reproducible_with_same_seed(self) -> None:
        sweep_a = sweep_asm2d_tsn_operating_space(n_samples=8, random_seed=5, show_progress=False)
        sweep_b = sweep_asm2d_tsn_operating_space(n_samples=8, random_seed=5, show_progress=False)

        pd = __import__("pandas")
        pd.testing.assert_frame_equal(sweep_a["influent_states"], sweep_b["influent_states"])
        pd.testing.assert_frame_equal(sweep_a["operating_conditions"], sweep_b["operating_conditions"])

    def test_sweep_samples_within_configured_ranges(self) -> None:
        params = load_asm2d_tsn_simulation_params()
        sweep_result = sweep_asm2d_tsn_operating_space(
            model_params=params, n_samples=16, random_seed=7, show_progress=False
        )
        state_columns = list(params["workbook"]["state_columns"])
        operational_columns = list(params["operational_columns"])
        influent_arr = sweep_result["influent_states"].to_numpy()
        operating_arr = sweep_result["operating_conditions"].to_numpy()

        for col_idx, state_name in enumerate(state_columns):
            lo, hi = params["influent_state_ranges"][state_name]
            self.assertTrue(np.all(influent_arr[:, col_idx] >= float(lo) - 1e-9), msg=f"{state_name} below lower bound")
            self.assertTrue(np.all(influent_arr[:, col_idx] <= float(hi) + 1e-9), msg=f"{state_name} above upper bound")

        for col_idx, op_name in enumerate(operational_columns):
            lo, hi = params["operational_ranges"][op_name]
            self.assertTrue(np.all(operating_arr[:, col_idx] >= float(lo) - 1e-9), msg=f"{op_name} below lower bound")
            self.assertTrue(np.all(operating_arr[:, col_idx] <= float(hi) + 1e-9), msg=f"{op_name} above upper bound")


if __name__ == "__main__":
    unittest.main()

