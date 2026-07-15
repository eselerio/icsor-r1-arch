"""ASM2D-TSN workbook, matrix, and mechanistic steady-state simulation helpers."""

from __future__ import annotations

import os
import re
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Mapping

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.integrate import solve_ivp
from scipy.optimize import least_squares
from tqdm import tqdm

from src.utils.io import (
    compute_file_sha256,
    load_json_file,
    load_pickle_file,
    save_json_file,
    save_pickle_file,
)
from src.utils.simulation import (
    get_repo_root,
    load_model_params,
    load_paths_config,
    render_simulation_artifact_paths,
    save_simulation_artifacts,
)


MODEL_NAME = "asm2d_tsn_simulation"
WORKBOOK_PATH_KEY = "asm2d_tsn_reference_workbook"
DATA_PATTERN_KEY = "asm2d_tsn_simulation_data_pattern"
METADATA_PATTERN_KEY = "asm2d_tsn_simulation_metadata_pattern"
COMPOSITION_CACHE_PATTERN_KEY = "asm2d_tsn_composition_cache_pattern"
COMPOSITION_CACHE_METADATA_PATTERN_KEY = "asm2d_tsn_composition_cache_metadata_pattern"
STOICHIOMETRIC_SHEET_NAME = "stoichiometric_matrix"
COMPOSITION_SHEET_NAME = "composition_matrix"
PARAMETER_SHEET_NAME = "parameter_table"
PARAMETER_VALUE_COLUMN_INDEX = 5

_EXCEL_REFERENCE_PATTERN = re.compile(
    r"(?<![A-Za-z0-9_])(?:(?:'(?P<sheet_quoted>[^']+)')|(?P<sheet_unquoted>[A-Za-z_][A-Za-z0-9_]*))?!?\$?(?P<column>[A-Za-z]{1,3})\$?(?P<row>\d+)"
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D8DEE6")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EEF2F5")
HEADER_FONT = Font(bold=True, color="22303C")

STOICHIOMETRIC_COEFFICIENTS: list[dict[str, dict[str, str]]] = [
    {"coefficients": {"S_F": "1-{f_SI}", "S_I": "{f_SI}", "X_S": "-1"}},
    {"coefficients": {"S_F": "1-{f_SI}", "S_I": "{f_SI}", "X_S": "-1"}},
    {"coefficients": {"S_F": "1-{f_SI}", "S_I": "{f_SI}", "X_S": "-1"}},
    {"coefficients": {"S_F": "1-{f_SI}", "S_I": "{f_SI}", "X_S": "-1"}},
    {"coefficients": {"S_F": "-1/{Y_H}", "S_O": "1-1/{Y_H}", "X_H": "1"}},
    {"coefficients": {"S_A": "-1/{Y_H}", "S_O": "1-1/{Y_H}", "X_H": "1"}},
    {
        "coefficients": {
            "S_F": "-1/{Y_H}",
            "S_NO2": "(1-{Y_H})/((8/7)*{Y_H})",
            "S_NO3": "-((1-{Y_H})/((8/7)*{Y_H}))",
            "X_H": "1",
        }
    },
    {
        "coefficients": {
            "S_F": "-1/{Y_H}",
            "S_N2": "(1-{Y_H})/(1.72*{Y_H})",
            "S_NO2": "-((1-{Y_H})/(1.72*{Y_H}))",
            "X_H": "1",
        }
    },
    {
        "coefficients": {
            "S_A": "-1/{Y_H}",
            "S_NO2": "(1-{Y_H})/((8/7)*{Y_H})",
            "S_NO3": "-((1-{Y_H})/((8/7)*{Y_H}))",
            "X_H": "1",
        }
    },
    {
        "coefficients": {
            "S_A": "-1/{Y_H}",
            "S_N2": "(1-{Y_H})/(1.72*{Y_H})",
            "S_NO2": "-((1-{Y_H})/(1.72*{Y_H}))",
            "X_H": "1",
        }
    },
    {"coefficients": {"S_A": "1", "S_F": "-1"}},
    {"coefficients": {"X_I": "{f_XI}", "X_S": "1-{f_XI}", "X_H": "-1"}},
    {"coefficients": {"S_A": "-1", "X_PP": "-{Y_PO4}", "X_PHA": "1"}},
    {"coefficients": {"S_O": "-{Y_PHA}", "X_PP": "1", "X_PHA": "-{Y_PHA}"}},
    {
        "coefficients": {
            "S_NO2": "{Y_PHA}/(8/7)",
            "S_NO3": "-({Y_PHA}/(8/7))",
            "X_PP": "1",
            "X_PHA": "-{Y_PHA}",
        }
    },
    {
        "coefficients": {
            "S_N2": "{Y_PHA}/1.72",
            "S_NO2": "-({Y_PHA}/1.72)",
            "X_PP": "1",
            "X_PHA": "-{Y_PHA}",
        }
    },
    {"coefficients": {"S_O": "1-1/{Y_PAO}", "X_PAO": "1", "X_PHA": "-1/{Y_PAO}"}},
    {
        "coefficients": {
            "S_NO2": "(1-{Y_PAO})/((8/7)*{Y_PAO})",
            "S_NO3": "-((1-{Y_PAO})/((8/7)*{Y_PAO}))",
            "X_PAO": "1",
            "X_PHA": "-1/{Y_PAO}",
        }
    },
    {
        "coefficients": {
            "S_N2": "(1-{Y_PAO})/(1.72*{Y_PAO})",
            "S_NO2": "-((1-{Y_PAO})/(1.72*{Y_PAO}))",
            "X_PAO": "1",
            "X_PHA": "-1/{Y_PAO}",
        }
    },
    {"coefficients": {"X_I": "{f_XI}", "X_S": "1-{f_XI}", "X_PAO": "-1"}},
    {"coefficients": {"X_PP": "-1"}},
    {"coefficients": {"S_A": "1", "X_PHA": "-1"}},
    {"coefficients": {"S_NO2": "1/{Y_AOB}", "S_O": "-((3.43-{Y_AOB})/{Y_AOB})", "X_AOB": "1"}},
    {"coefficients": {"S_NO2": "-1/{Y_NOB}", "S_NO3": "1/{Y_NOB}", "S_O": "-((1.14-{Y_NOB})/{Y_NOB})", "X_NOB": "1"}},
    {"coefficients": {"X_I": "{f_XI}", "X_S": "1-{f_XI}", "X_AOB": "-1"}},
    {"coefficients": {"X_I": "{f_XI}", "X_S": "1-{f_XI}", "X_NOB": "-1"}},
    {"coefficients": {"S_PO4": "-1", "X_MeOH": "-3.45", "X_MeP": "4.87"}},
    {"coefficients": {"S_PO4": "1", "X_MeOH": "3.45", "X_MeP": "-4.87"}},
]

COMPOSITION_FORMULAS: dict[str, dict[str, str]] = {
    "S_A": {"COD": "1"},
    "S_F": {"COD": "1", "TN": "{i_NSF}", "TKN": "{i_NSF}", "TP": "{i_PSF}"},
    "S_I": {"COD": "1", "TN": "{i_NSI}", "TKN": "{i_NSI}", "TP": "{i_PSI}"},
    "S_NH4": {"TN": "1", "TKN": "1"},
    "S_NO2": {"TN": "1"},
    "S_NO3": {"TN": "1"},
    "S_PO4": {"TP": "1"},
    "X_I": {"COD": "1", "TN": "{i_NXI}", "TKN": "{i_NXI}", "TP": "{i_PXI}", "TSS": "{i_TSS_XI}"},
    "X_S": {"COD": "1", "TN": "{i_NXS}", "TKN": "{i_NXS}", "TP": "{i_PXS}", "TSS": "{i_TSS_XS}"},
    "X_H": {"COD": "1", "TN": "{i_NBM}", "TKN": "{i_NBM}", "TP": "{i_PBM}", "TSS": "{i_TSS_BM}"},
    "X_PAO": {"COD": "1", "TN": "{i_NBM}", "TKN": "{i_NBM}", "TP": "{i_PBM}", "TSS": "{i_TSS_BM}"},
    "X_PP": {"TP": "1", "TSS": "{i_TSS_PP}"},
    "X_PHA": {"COD": "1", "TSS": "{i_TSS_PHA}"},
    "X_AOB": {"COD": "1", "TN": "{i_NBM}", "TKN": "{i_NBM}", "TP": "{i_PBM}", "TSS": "{i_TSS_BM}"},
    "X_NOB": {"COD": "1", "TN": "{i_NBM}", "TKN": "{i_NBM}", "TP": "{i_PBM}", "TSS": "{i_TSS_BM}"},
    "X_MeOH": {"TSS": "1"},
    "X_MeP": {"TP": "{i_PMeP}", "TSS": "1"},
}

NITROGEN_CONTINUITY_TERMS = {
    "S_F": "{i_NSF}",
    "S_I": "{i_NSI}",
    "S_N2": "1",
    "S_NO2": "1",
    "S_NO3": "1",
    "X_I": "{i_NXI}",
    "X_S": "{i_NXS}",
    "X_H": "{i_NBM}",
    "X_PAO": "{i_NBM}",
    "X_AOB": "{i_NBM}",
    "X_NOB": "{i_NBM}",
}

PHOSPHORUS_CONTINUITY_TERMS = {
    "S_F": "{i_PSF}",
    "S_I": "{i_PSI}",
    "X_I": "{i_PXI}",
    "X_S": "{i_PXS}",
    "X_H": "{i_PBM}",
    "X_PAO": "{i_PBM}",
    "X_PP": "1",
    "X_AOB": "{i_PBM}",
    "X_NOB": "{i_PBM}",
    "X_MeP": "{i_PMeP}",
}

def load_asm2d_tsn_simulation_params(repo_root: str | Path | None = None) -> dict[str, Any]:
    """Load the configured ASM2D-TSN simulation definition."""

    return load_model_params(MODEL_NAME, repo_root)


def resolve_asm2d_tsn_workbook_path(
    repo_root: str | Path | None = None,
    *,
    paths_config: Mapping[str, Any] | None = None,
) -> Path:
    """Resolve the configured canonical workbook path."""

    root = get_repo_root(repo_root)
    config = dict(paths_config) if paths_config is not None else load_paths_config(root)
    return root / Path(config[WORKBOOK_PATH_KEY])


def resolve_asm2d_tsn_simulation_artifact_paths(
    repo_root: str | Path | None = None,
    *,
    timestamp: str | None = None,
    paths_config: Mapping[str, Any] | None = None,
) -> tuple[Path, Path, str]:
    """Resolve the configured ASM2D-TSN dataset and metadata output paths."""

    return render_simulation_artifact_paths(
        MODEL_NAME,
        repo_root=repo_root,
        timestamp=timestamp,
        paths_config=paths_config,
        data_pattern_key=DATA_PATTERN_KEY,
        metadata_pattern_key=METADATA_PATTERN_KEY,
    )


def resolve_asm2d_tsn_composition_cache_paths(
    workbook_hash: str,
    repo_root: str | Path | None = None,
    *,
    paths_config: Mapping[str, Any] | None = None,
) -> tuple[Path, Path]:
    """Resolve configured cache paths for one workbook-derived composition artifact."""

    root = get_repo_root(repo_root)
    config = dict(paths_config) if paths_config is not None else load_paths_config(root)
    matrix_path = root / Path(config[COMPOSITION_CACHE_PATTERN_KEY].format(workbook_hash=str(workbook_hash)))
    metadata_path = root / Path(
        config[COMPOSITION_CACHE_METADATA_PATTERN_KEY].format(workbook_hash=str(workbook_hash))
    )
    return matrix_path, metadata_path


def create_asm2d_tsn_workbook(
    workbook_path: str | Path | None = None,
    *,
    repo_root: str | Path | None = None,
    model_params: Mapping[str, Any] | None = None,
) -> Path:
    """Create the canonical ASM2D-TSN workbook with formula-driven matrices."""

    workbook = build_asm2d_tsn_workbook(model_params=model_params, repo_root=repo_root)
    output_path = Path(workbook_path) if workbook_path is not None else resolve_asm2d_tsn_workbook_path(repo_root)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path.resolve()


def build_asm2d_tsn_workbook(
    *,
    model_params: Mapping[str, Any] | None = None,
    repo_root: str | Path | None = None,
) -> Workbook:
    """Build the workbook object for the configured ASM2D-TSN reference model."""

    params = dict(model_params) if model_params is not None else load_asm2d_tsn_simulation_params(repo_root)
    workbook_config = _validate_workbook_config(params)
    parameter_refs = _build_parameter_reference_map(workbook_config["parameters"])

    workbook = Workbook()
    stoichiometric_sheet = workbook.active
    stoichiometric_sheet.title = STOICHIOMETRIC_SHEET_NAME
    composition_sheet = workbook.create_sheet(COMPOSITION_SHEET_NAME)
    parameter_sheet = workbook.create_sheet(PARAMETER_SHEET_NAME)

    _write_stoichiometric_sheet(stoichiometric_sheet, workbook_config, parameter_refs)
    _write_composition_sheet(composition_sheet, workbook_config, parameter_refs)
    _write_parameter_sheet(parameter_sheet, workbook_config["parameters"])

    for worksheet in workbook.worksheets:
        _auto_size_columns(worksheet)
        worksheet.auto_filter.ref = worksheet.dimensions

    return workbook


def _normalize_excel_reference(sheet_name: str, coordinate: str) -> str:
    return f"{str(sheet_name).strip().lower()}!{str(coordinate).replace('$', '').upper()}"


def _build_workbook_numeric_lookup(workbook) -> dict[str, float]:
    numeric_lookup: dict[str, float] = {}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell_value = cell.value
                if isinstance(cell_value, bool):
                    continue
                if isinstance(cell_value, (int, float)):
                    numeric_lookup[_normalize_excel_reference(worksheet.title, cell.coordinate)] = float(cell_value)
    return numeric_lookup


def _evaluate_workbook_formula(
    formula: str,
    numeric_lookup: Mapping[str, float],
    *,
    current_sheet_name: str,
) -> float:
    expression = str(formula).strip()
    if expression.startswith("="):
        expression = expression[1:]

    def _replace_reference(match: re.Match[str]) -> str:
        sheet_name = match.group("sheet_quoted") or match.group("sheet_unquoted") or str(current_sheet_name)
        coordinate = f"{match.group('column').upper()}{match.group('row')}"
        lookup_key = _normalize_excel_reference(sheet_name, coordinate)
        if lookup_key not in numeric_lookup:
            raise KeyError(
                "Workbook formula references a non-numeric or missing cell: "
                f"{sheet_name}!{coordinate}."
            )
        return str(float(numeric_lookup[lookup_key]))

    resolved_expression = _EXCEL_REFERENCE_PATTERN.sub(_replace_reference, expression).replace("^", "**")
    return float(eval(resolved_expression, {"__builtins__": {}}, {}))


def _coerce_workbook_composition_value(
    cell_value: Any,
    numeric_lookup: Mapping[str, float],
    *,
    current_sheet_name: str,
) -> float:
    if cell_value is None:
        return 0.0
    if isinstance(cell_value, bool):
        return float(cell_value)
    if isinstance(cell_value, (int, float)):
        return float(cell_value)

    text_value = str(cell_value).strip()
    if not text_value:
        return 0.0
    if text_value.startswith("="):
        return _evaluate_workbook_formula(
            text_value,
            numeric_lookup,
            current_sheet_name=current_sheet_name,
        )

    try:
        return float(text_value)
    except ValueError as error:
        raise ValueError(f"Workbook composition_matrix contains a non-numeric value: {text_value!r}") from error


def _read_composition_matrix_from_workbook(
    workbook_path: str | Path,
    *,
    expected_state_columns: list[str],
) -> dict[str, Any]:
    workbook = load_workbook(filename=Path(workbook_path), data_only=False)
    try:
        if COMPOSITION_SHEET_NAME not in workbook.sheetnames:
            raise KeyError(f"Workbook is missing required sheet '{COMPOSITION_SHEET_NAME}'.")
        if PARAMETER_SHEET_NAME not in workbook.sheetnames:
            raise KeyError(f"Workbook is missing required sheet '{PARAMETER_SHEET_NAME}'.")

        worksheet = workbook[COMPOSITION_SHEET_NAME]
        header_values = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
        if "state_variable" not in header_values:
            raise KeyError("Workbook composition_matrix must include a 'state_variable' header column.")

        state_column_number = header_values.index("state_variable") + 1
        reserved_headers = {"state_group", "state_variable", "unit"}
        composite_header_pairs = [
            (column_number, header_name)
            for column_number, header_name in enumerate(header_values, start=1)
            if header_name and header_name not in reserved_headers
        ]
        if not composite_header_pairs:
            raise ValueError("Workbook composition_matrix must define at least one composite output column.")

        measured_output_columns = [header_name for _, header_name in composite_header_pairs]
        _validate_unique_names(measured_output_columns, "composition_matrix output columns")

        numeric_lookup = _build_workbook_numeric_lookup(workbook)
        state_columns: list[str] = []
        coefficients_by_state: list[list[float]] = []
        for row_number in range(2, worksheet.max_row + 1):
            raw_state_name = worksheet.cell(row=row_number, column=state_column_number).value
            if raw_state_name is None:
                continue
            state_name = str(raw_state_name).strip()
            if not state_name:
                continue

            state_columns.append(state_name)
            row_coefficients: list[float] = []
            for column_number, _ in composite_header_pairs:
                raw_value = worksheet.cell(row=row_number, column=column_number).value
                row_coefficients.append(
                    _coerce_workbook_composition_value(
                        raw_value,
                        numeric_lookup,
                        current_sheet_name=worksheet.title,
                    )
                )
            coefficients_by_state.append(row_coefficients)

        if not state_columns:
            raise ValueError("Workbook composition_matrix must define at least one state_variable row.")
        _validate_unique_names(state_columns, "composition_matrix state_variable")

        if state_columns != expected_state_columns:
            raise ValueError(
                "Workbook composition_matrix state_variable rows must match configured workbook state_columns "
                "exactly and in order."
            )

        composition_matrix = np.asarray(coefficients_by_state, dtype=float).T
        expected_shape = (len(measured_output_columns), len(state_columns))
        if composition_matrix.shape != expected_shape:
            raise ValueError(
                "Workbook composition_matrix shape must match measured_output_columns x state_columns."
            )

        return {
            "state_columns": state_columns,
            "measured_output_columns": measured_output_columns,
            "composition_matrix": composition_matrix,
        }
    finally:
        workbook.close()


def _build_workbook_fingerprint(workbook_path: Path) -> dict[str, Any]:
    resolved_path = Path(workbook_path).resolve()
    stat_info = resolved_path.stat()
    return {
        "workbook_path": resolved_path.as_posix(),
        "workbook_sha256": compute_file_sha256(resolved_path),
        "workbook_mtime_ns": int(stat_info.st_mtime_ns),
        "workbook_size_bytes": int(stat_info.st_size),
    }


def _validate_cached_composition_payload(cache_payload: Mapping[str, Any]) -> None:
    required_keys = ("state_columns", "measured_output_columns", "composition_matrix")
    for required_key in required_keys:
        if required_key not in cache_payload:
            raise KeyError(f"Cached composition payload is missing required key '{required_key}'.")


def load_asm2d_tsn_workbook_composition(
    *,
    repo_root: str | Path | None = None,
    workbook_path: str | Path | None = None,
    model_params: Mapping[str, Any] | None = None,
    paths_config: Mapping[str, Any] | None = None,
    use_cache: bool = True,
) -> dict[str, Any]:
    """Load workbook-derived composition schema and coefficients, with fingerprinted cache reuse."""

    params = dict(model_params) if model_params is not None else load_asm2d_tsn_simulation_params(repo_root)
    workbook_config = _validate_workbook_config(params)
    expected_state_columns = list(workbook_config["state_columns"])
    resolved_workbook_path = (
        Path(workbook_path)
        if workbook_path is not None
        else resolve_asm2d_tsn_workbook_path(repo_root, paths_config=paths_config)
    )
    if not resolved_workbook_path.exists():
        raise FileNotFoundError(f"ASM2D-TSN workbook not found at {resolved_workbook_path}.")

    fingerprint = _build_workbook_fingerprint(resolved_workbook_path)
    cache_matrix_path, cache_metadata_path = resolve_asm2d_tsn_composition_cache_paths(
        fingerprint["workbook_sha256"],
        repo_root=repo_root,
        paths_config=paths_config,
    )

    if use_cache and cache_matrix_path.exists() and cache_metadata_path.exists():
        cache_metadata = load_json_file(cache_metadata_path)
        if (
            str(cache_metadata.get("workbook_sha256")) == str(fingerprint["workbook_sha256"])
            and int(cache_metadata.get("workbook_mtime_ns", -1)) == int(fingerprint["workbook_mtime_ns"])
            and int(cache_metadata.get("workbook_size_bytes", -1)) == int(fingerprint["workbook_size_bytes"])
        ):
            cache_payload = load_pickle_file(cache_matrix_path)
            _validate_cached_composition_payload(cache_payload)
            state_columns = [str(name) for name in cache_payload["state_columns"]]
            measured_output_columns = [str(name) for name in cache_payload["measured_output_columns"]]
            if state_columns != expected_state_columns:
                raise ValueError(
                    "Cached composition state_columns no longer match configured workbook state_columns."
                )
            composition_matrix = np.asarray(cache_payload["composition_matrix"], dtype=float)
            if composition_matrix.shape != (len(measured_output_columns), len(state_columns)):
                raise ValueError("Cached composition_matrix shape is invalid for cached schema.")
            return {
                "state_columns": state_columns,
                "measured_output_columns": measured_output_columns,
                "composition_matrix": composition_matrix,
                **fingerprint,
                "cache_source": "cache",
                "cache_paths": {
                    "composition_matrix": cache_matrix_path,
                    "composition_metadata": cache_metadata_path,
                },
            }

    parsed_composition = _read_composition_matrix_from_workbook(
        resolved_workbook_path,
        expected_state_columns=expected_state_columns,
    )
    composition_payload = {
        "state_columns": list(parsed_composition["state_columns"]),
        "measured_output_columns": list(parsed_composition["measured_output_columns"]),
        "composition_matrix": np.asarray(parsed_composition["composition_matrix"], dtype=float),
    }

    if use_cache:
        save_pickle_file(cache_matrix_path, composition_payload)
        save_json_file(
            cache_metadata_path,
            {
                "cache_schema_version": 1,
                "state_columns": composition_payload["state_columns"],
                "measured_output_columns": composition_payload["measured_output_columns"],
                **fingerprint,
            },
        )

    return {
        **composition_payload,
        **fingerprint,
        "cache_source": "workbook",
        "cache_paths": {
            "composition_matrix": cache_matrix_path,
            "composition_metadata": cache_metadata_path,
        },
    }


def get_asm2d_tsn_matrices(
    model_params: Mapping[str, Any] | None = None,
    *,
    repo_root: str | Path | None = None,
    paths_config: Mapping[str, Any] | None = None,
    use_composition_cache: bool = True,
) -> dict[str, Any]:
    """Build numeric Petersen and composition matrices for the configured ASM2D-TSN model."""

    params = dict(model_params) if model_params is not None else load_asm2d_tsn_simulation_params(repo_root)
    composition_bundle = load_asm2d_tsn_workbook_composition(
        repo_root=repo_root,
        model_params=params,
        paths_config=paths_config,
        use_cache=use_composition_cache,
    )
    measured_output_columns = list(composition_bundle["measured_output_columns"])
    runtime = _validate_runtime_structure(params, measured_output_columns=measured_output_columns)
    workbook_config = runtime["workbook_config"]
    parameter_values = _build_parameter_value_map(workbook_config["parameters"])
    state_columns = list(runtime["state_columns"])
    process_names = list(runtime["process_names"])
    process_types = list(runtime["process_types"])
    state_index = _build_state_index(state_columns)

    composition_state_columns = list(composition_bundle["state_columns"])
    if composition_state_columns != state_columns:
        raise ValueError(
            "Workbook composition_matrix state columns do not match the configured ASM2D-TSN state columns."
        )

    petersen_matrix = np.zeros((len(process_names), len(state_columns)), dtype=float)
    composition_matrix = np.asarray(composition_bundle["composition_matrix"], dtype=float)
    if composition_matrix.shape != (len(measured_output_columns), len(state_columns)):
        raise ValueError(
            "Workbook composition_matrix shape must match measured_output_columns x state_columns."
        )

    for row_index, process_definition in enumerate(STOICHIOMETRIC_COEFFICIENTS):
        row_values = petersen_matrix[row_index]
        direct_coefficients = process_definition["coefficients"]

        for state_name, expression in direct_coefficients.items():
            row_values[state_index[state_name]] = _evaluate_numeric_expression(expression, parameter_values)

        row_values[state_index["S_NH4"]] = -sum(
            row_values[state_index[state_name]] * _evaluate_numeric_expression(factor_expression, parameter_values)
            for state_name, factor_expression in NITROGEN_CONTINUITY_TERMS.items()
        )

        if "S_PO4" not in direct_coefficients:
            row_values[state_index["S_PO4"]] = -sum(
                row_values[state_index[state_name]] * _evaluate_numeric_expression(factor_expression, parameter_values)
                for state_name, factor_expression in PHOSPHORUS_CONTINUITY_TERMS.items()
            )

        row_values[state_index["S_ALK"]] = (
            row_values[state_index["S_NH4"]] / 14.0
            - row_values[state_index["S_NO2"]] / 14.0
            - row_values[state_index["S_NO3"]] / 14.0
            + row_values[state_index["S_PO4"]] / 31.0
        )

    return {
        "petersen_matrix": petersen_matrix,
        "composition_matrix": composition_matrix,
        "process_names": process_names,
        "process_types": process_types,
        "state_index": state_index,
        "state_columns": state_columns,
        "measured_output_columns": measured_output_columns,
        "composition_workbook_path": str(composition_bundle["workbook_path"]),
        "composition_workbook_sha256": str(composition_bundle["workbook_sha256"]),
        "composition_workbook_mtime_ns": int(composition_bundle["workbook_mtime_ns"]),
        "composition_workbook_size_bytes": int(composition_bundle["workbook_size_bytes"]),
        "composition_cache_source": str(composition_bundle["cache_source"]),
        "composition_cache_paths": dict(composition_bundle["cache_paths"]),
    }


def build_asm2d_tsn_metadata(
    model_params: Mapping[str, Any],
    *,
    sample_count: int,
    random_seed: int,
    dataset_file: str | None = None,
    measured_output_columns: list[str] | None = None,
    composition_source: Mapping[str, Any] | None = None,
    repo_root: str | Path | None = None,
) -> dict[str, Any]:
    """Create the metadata contract for the ASM2D-TSN mixed-schema dataset."""

    if measured_output_columns is None:
        measured_output_columns = list(
            load_asm2d_tsn_workbook_composition(
                repo_root=repo_root,
                model_params=model_params,
            )["measured_output_columns"]
        )

    runtime = _validate_runtime_structure(model_params, measured_output_columns=measured_output_columns)
    state_columns = list(runtime["state_columns"])
    measured_output_columns = list(runtime["measured_output_columns"])
    process_names = list(runtime["process_names"])
    process_types = list(runtime["process_types"])
    operational_columns = list(runtime["operational_columns"])
    influent_fraction_columns = [f"In_{name}" for name in state_columns]
    influent_composite_columns = [f"In_{name}" for name in measured_output_columns]
    effluent_fraction_columns = [f"Out_{name}" for name in state_columns]
    dependent_columns = [f"Out_{name}" for name in measured_output_columns]

    return {
        "simulation_name": MODEL_NAME,
        "n_samples": sample_count,
        "random_seed": random_seed,
        "sampling_method": "latin_hypercube",
        "dependent_columns": dependent_columns,
        "independent_columns": operational_columns + influent_fraction_columns,
        "identifier_columns": [],
        "ignored_columns": influent_composite_columns + effluent_fraction_columns,
        "dataset_file": dataset_file,
        "state_columns": state_columns,
        "measured_output_columns": measured_output_columns,
        "operational_columns": operational_columns,
        "influent_fraction_columns": influent_fraction_columns,
        "influent_composite_columns": influent_composite_columns,
        "effluent_fraction_columns": effluent_fraction_columns,
        "effluent_composite_columns": dependent_columns,
        "processes": process_names,
        "process_types": process_types,
        "petersen_matrix_shape": [len(process_names), len(state_columns)],
        "composition_matrix_shape": [len(measured_output_columns), len(state_columns)],
        "schema_version": str(model_params["schema_version"]),
        "composition_source": dict(composition_source or {}),
    }


def generate_asm2d_tsn_dataset(
    *,
    model_params: Mapping[str, Any] | None = None,
    repo_root: str | Path | None = None,
    n_samples: int | None = None,
    random_seed: int | None = None,
    parallel_workers: int | None = None,
    parallel_chunk_size: int | None = None,
    include_debug_data: bool = False,
    show_progress: bool = False,
    progress_description: str | None = None,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]]:
    """Generate a mechanistic steady-state ASM2D-TSN dataset with input/output fractions and composites."""

    params = dict(model_params) if model_params is not None else load_asm2d_tsn_simulation_params(repo_root)
    matrix_bundle = get_asm2d_tsn_matrices(params, repo_root=repo_root)
    runtime = _validate_runtime_structure(
        params,
        measured_output_columns=list(matrix_bundle["measured_output_columns"]),
    )
    configured_hyperparameters = params["hyperparameters"]
    sample_count = int(n_samples if n_samples is not None else configured_hyperparameters["n_samples"])
    if sample_count < 0:
        raise ValueError("n_samples must be greater than or equal to 0.")

    seed = int(random_seed if random_seed is not None else configured_hyperparameters["seed"])
    requested_parallel_workers = int(
        parallel_workers if parallel_workers is not None else configured_hyperparameters.get("parallel_workers", 1)
    )
    requested_parallel_chunk_size = int(
        parallel_chunk_size
        if parallel_chunk_size is not None
        else configured_hyperparameters.get("parallel_chunk_size", sample_count or 1)
    )
    state_columns = list(runtime["state_columns"])
    operational_columns = list(runtime["operational_columns"])
    measured_output_columns = list(runtime["measured_output_columns"])

    influent_states = np.zeros((sample_count, len(state_columns)), dtype=float)
    operational = np.zeros((sample_count, len(operational_columns)), dtype=float)
    measured_outputs = np.zeros((sample_count, len(measured_output_columns)), dtype=float)
    effluent_states = np.zeros((sample_count, len(state_columns)), dtype=float)
    solver_diagnostic_records: list[dict[str, Any]] = []
    chunk_results: list[tuple[int, np.ndarray, np.ndarray, np.ndarray, np.ndarray, list[dict[str, Any]]]] = []
    if sample_count > 0:
        worker_count = _resolve_parallel_workers(requested_parallel_workers, sample_count)
        chunk_size = min(_resolve_parallel_chunk_size(requested_parallel_chunk_size), sample_count)
        chunk_specs = [
            {
                "chunk_start": chunk_start,
                "chunk_size": min(chunk_size, sample_count - chunk_start),
                "chunk_seed": seed + chunk_index,
                "model_params": params,
                "matrix_bundle": matrix_bundle,
                "runtime": runtime,
                "collect_debug_data": include_debug_data,
            }
            for chunk_index, chunk_start in enumerate(range(0, sample_count, chunk_size))
        ]

        if worker_count == 1 or len(chunk_specs) == 1:
            progress_bar = tqdm(
                total=sample_count,
                desc=progress_description or "ASM2D-TSN simulation",
                unit="sample",
                disable=not show_progress,
            )
            try:
                chunk_results = [
                    _generate_asm2d_tsn_dataset_chunk(progress_bar=progress_bar, **chunk_spec)
                    for chunk_spec in chunk_specs
                ]
            finally:
                progress_bar.close()
        else:
            with ProcessPoolExecutor(max_workers=worker_count) as executor:
                future_map = {
                    executor.submit(_generate_asm2d_tsn_dataset_chunk, **chunk_spec): int(chunk_spec["chunk_size"])
                    for chunk_spec in chunk_specs
                }
                progress_bar = tqdm(
                    total=sample_count,
                    desc=progress_description or "ASM2D-TSN simulation",
                    unit="sample",
                    disable=not show_progress,
                )
                try:
                    for future in as_completed(future_map):
                        chunk_results.append(future.result())
                        progress_bar.update(future_map[future])
                finally:
                    progress_bar.close()

    for chunk_start, chunk_influent, chunk_operational, chunk_effluent, chunk_measured, chunk_diagnostics in chunk_results:
        chunk_end = chunk_start + len(chunk_influent)
        influent_states[chunk_start:chunk_end] = chunk_influent
        operational[chunk_start:chunk_end] = chunk_operational
        effluent_states[chunk_start:chunk_end] = chunk_effluent
        measured_outputs[chunk_start:chunk_end] = chunk_measured
        if include_debug_data:
            solver_diagnostic_records.extend(chunk_diagnostics)

    influent_df = pd.DataFrame(influent_states, columns=[f"In_{name}" for name in state_columns])
    influent_composite_df = pd.DataFrame(
        influent_states @ np.asarray(matrix_bundle["composition_matrix"], dtype=float).T,
        columns=[f"In_{name}" for name in measured_output_columns],
    )
    operational_df = pd.DataFrame(operational, columns=operational_columns)
    effluent_df = pd.DataFrame(effluent_states, columns=[f"Out_{name}" for name in state_columns])
    measured_df = pd.DataFrame(measured_outputs, columns=[f"Out_{name}" for name in measured_output_columns])
    dataset = pd.concat([operational_df, influent_df, influent_composite_df, effluent_df, measured_df], axis=1)

    metadata = build_asm2d_tsn_metadata(
        params,
        sample_count=sample_count,
        random_seed=seed,
        measured_output_columns=measured_output_columns,
        composition_source={
            "workbook_path": matrix_bundle["composition_workbook_path"],
            "workbook_sha256": matrix_bundle["composition_workbook_sha256"],
            "workbook_mtime_ns": matrix_bundle["composition_workbook_mtime_ns"],
            "workbook_size_bytes": matrix_bundle["composition_workbook_size_bytes"],
            "cache_source": matrix_bundle["composition_cache_source"],
        },
        repo_root=repo_root,
    )

    simulation_bundle = dict(matrix_bundle)
    if include_debug_data:
        solver_diagnostics = pd.DataFrame(solver_diagnostic_records).sort_values("sample_index").reset_index(drop=True)
        simulation_bundle["effluent_states"] = pd.DataFrame(effluent_states, columns=state_columns)
        simulation_bundle["solver_diagnostics"] = solver_diagnostics
        simulation_bundle["solver_summary"] = _summarize_asm2d_tsn_solver_diagnostics(
            solver_diagnostics,
            float(runtime["solver"]["acceptance_residual_max"]),
        )

    return dataset, metadata, simulation_bundle


def run_asm2d_tsn_simulation(
    *,
    save_artifacts: bool = True,
    repo_root: str | Path | None = None,
    n_samples: int | None = None,
    random_seed: int | None = None,
    parallel_workers: int | None = None,
    parallel_chunk_size: int | None = None,
    timestamp: str | None = None,
    include_debug_data: bool = False,
    show_progress: bool = False,
    progress_description: str | None = None,
) -> dict[str, Any]:
    """Run the ASM2D-TSN steady-state simulation and optionally persist artifacts."""

    params = load_asm2d_tsn_simulation_params(repo_root)
    simulation_bundle: dict[str, Any]
    dataset, metadata, simulation_bundle = generate_asm2d_tsn_dataset(
        model_params=params,
        repo_root=repo_root,
        n_samples=n_samples,
        random_seed=random_seed,
        parallel_workers=parallel_workers,
        parallel_chunk_size=parallel_chunk_size,
        include_debug_data=include_debug_data,
        show_progress=show_progress,
        progress_description=progress_description,
    )

    artifact_paths: dict[str, Path | None] = {
        "dataset_csv": None,
        "metadata_json": None,
    }

    if save_artifacts:
        dataset_path, metadata_path, persisted_metadata = save_simulation_artifacts(
            dataset,
            metadata,
            MODEL_NAME,
            repo_root=repo_root,
            timestamp=timestamp,
            data_pattern_key=DATA_PATTERN_KEY,
            metadata_pattern_key=METADATA_PATTERN_KEY,
        )
        metadata = persisted_metadata
        artifact_paths = {
            "dataset_csv": dataset_path,
            "metadata_json": metadata_path,
        }

    return {
        "dataset": dataset,
        "metadata": metadata,
        "petersen_matrix": simulation_bundle["petersen_matrix"],
        "composition_matrix": simulation_bundle["composition_matrix"],
        "matrix_bundle": simulation_bundle,
        "composite_matrix": simulation_bundle["composition_matrix"],
        "artifact_paths": artifact_paths,
        "effluent_states": simulation_bundle.get("effluent_states"),
        "solver_diagnostics": simulation_bundle.get("solver_diagnostics"),
        "solver_summary": simulation_bundle.get("solver_summary"),
    }


def sweep_asm2d_tsn_operating_space(
    *,
    model_params: Mapping[str, Any] | None = None,
    repo_root: str | Path | None = None,
    n_samples: int = 512,
    random_seed: int | None = None,
    show_progress: bool = False,
    progress_description: str | None = None,
) -> dict[str, Any]:
    """Sample the configured operating space and summarize ASM2D-TSN solver behavior."""

    if n_samples < 1:
        raise ValueError("n_samples must be at least 1.")

    params = dict(model_params) if model_params is not None else load_asm2d_tsn_simulation_params(repo_root)
    matrix_bundle = get_asm2d_tsn_matrices(params, repo_root=repo_root)
    runtime = _validate_runtime_structure(
        params,
        measured_output_columns=list(matrix_bundle["measured_output_columns"]),
    )
    configured_hyperparameters = params["hyperparameters"]
    seed = int(random_seed if random_seed is not None else configured_hyperparameters["seed"])
    state_columns = list(runtime["state_columns"])
    operational_columns = list(runtime["operational_columns"])
    state_index = dict(matrix_bundle["state_index"])

    # Pre-generate a joint LHS design for all sweep points.
    all_columns = state_columns + operational_columns
    all_ranges = {**runtime["influent_state_ranges"], **runtime["operational_ranges"]}
    candidate_pool = _generate_lhs_candidate_pool(seed, n_samples, all_columns, all_ranges)
    n_state = len(state_columns)

    influent_samples = np.zeros((n_samples, len(state_columns)), dtype=float)
    operational_samples = np.zeros((n_samples, len(operational_columns)), dtype=float)
    effluent_samples = np.zeros((n_samples, len(state_columns)), dtype=float)
    diagnostic_records: list[dict[str, Any]] = []

    progress_bar = tqdm(
        total=n_samples,
        desc=progress_description or "ASM2D-TSN sweep",
        unit="sample",
        disable=not show_progress,
    )
    try:
        for sample_index in range(n_samples):
            sampled_state = candidate_pool[sample_index, :n_state]
            influent_state = _build_influent_state_sample(sampled_state)
            operating_point = candidate_pool[sample_index, n_state:]
            effluent_state, diagnostics = simulate_asm2d_tsn_steady_state(
                influent_state=influent_state,
                hrt_hours=float(operating_point[0]),
                aeration=float(operating_point[1]),
                model_params=params,
                matrix_bundle=matrix_bundle,
                previous_solution=None,
                enforce_acceptance=False,
            )
            influent_samples[sample_index] = influent_state
            operational_samples[sample_index] = operating_point
            effluent_samples[sample_index] = effluent_state
            diagnostic_record = dict(diagnostics)
            diagnostic_record["sample_index"] = sample_index
            diagnostic_record["HRT"] = float(operating_point[0])
            diagnostic_record["Aeration"] = float(operating_point[1])
            diagnostic_records.append(diagnostic_record)
            progress_bar.update(1)
    finally:
        progress_bar.close()

    solver_diagnostics = pd.DataFrame(diagnostic_records).sort_values("sample_index").reset_index(drop=True)
    summary = _summarize_asm2d_tsn_solver_diagnostics(
        solver_diagnostics,
        float(runtime["solver"]["acceptance_residual_max"]),
    )

    return {
        "influent_states": pd.DataFrame(influent_samples, columns=state_columns),
        "operating_conditions": pd.DataFrame(operational_samples, columns=operational_columns),
        "effluent_states": pd.DataFrame(effluent_samples, columns=state_columns),
        "solver_diagnostics": solver_diagnostics,
        "summary": summary,
        "matrix_bundle": matrix_bundle,
    }


def _validate_workbook_config(model_params: Mapping[str, Any]) -> dict[str, Any]:
    if "workbook" not in model_params:
        raise KeyError("asm2d_tsn_simulation must define a workbook section.")

    workbook_config = dict(model_params["workbook"])
    expected_sheets = [STOICHIOMETRIC_SHEET_NAME, COMPOSITION_SHEET_NAME, PARAMETER_SHEET_NAME]
    configured_sheets = list(workbook_config["sheets"])
    if configured_sheets != expected_sheets:
        raise ValueError("asm2d_tsn_simulation workbook sheets must match the required three-sheet contract.")

    dissolved_state_columns = list(workbook_config["dissolved_state_columns"])
    particulate_state_columns = list(workbook_config["particulate_state_columns"])
    state_columns = list(workbook_config["state_columns"])
    legacy_composite_variables = workbook_config.get("composite_variables")
    processes = list(workbook_config["processes"])
    parameter_rows = list(workbook_config["parameters"])
    state_units = dict(workbook_config["state_units"])

    _validate_unique_names(dissolved_state_columns, "dissolved_state_columns")
    _validate_unique_names(particulate_state_columns, "particulate_state_columns")
    _validate_unique_names(state_columns, "state_columns")
    _resolve_workbook_composite_variables(
        state_columns,
        legacy_composite_variables=(
            None if legacy_composite_variables is None else list(legacy_composite_variables)
        ),
    )

    if state_columns != dissolved_state_columns + particulate_state_columns:
        raise ValueError("asm2d_tsn_simulation state_columns must concatenate dissolved and particulate state columns.")

    missing_state_units = [state_name for state_name in state_columns if state_name not in state_units]
    if missing_state_units:
        missing_display = ", ".join(missing_state_units)
        raise KeyError(f"asm2d_tsn_simulation missing state units for: {missing_display}")

    if len(processes) != len(STOICHIOMETRIC_COEFFICIENTS):
        raise ValueError("asm2d_tsn_simulation workbook process count does not match the stoichiometric matrix definition.")

    process_indices = [int(process["index"]) for process in processes]
    if process_indices != list(range(1, len(processes) + 1)):
        raise ValueError("asm2d_tsn_simulation workbook processes must be sequentially indexed from 1.")

    parameter_names = [str(parameter_row["excel_name"]) for parameter_row in parameter_rows]
    _validate_unique_names(parameter_names, "parameter excel_name")

    for parameter_row in parameter_rows:
        for required_key in ("category", "symbol", "excel_name", "description", "value", "unit"):
            if required_key not in parameter_row:
                raise KeyError(f"asm2d_tsn_simulation parameter row missing '{required_key}'.")
        float(parameter_row["value"])

    return workbook_config


def _validate_runtime_structure(
    model_params: Mapping[str, Any],
    *,
    measured_output_columns: list[str] | None,
) -> dict[str, Any]:
    workbook_config = _validate_workbook_config(model_params)
    solver = _validate_solver_config(model_params)
    state_columns = list(workbook_config["state_columns"])
    if measured_output_columns is None:
        raise ValueError(
            "asm2d_tsn_simulation requires measured_output_columns derived from workbook composition_matrix."
        )
    measured_output_columns = [str(name) for name in measured_output_columns]
    process_names = [str(process["name"]) for process in workbook_config["processes"]]
    process_types = list(model_params["process_types"])
    operational_columns = list(model_params["operational_columns"])
    influent_state_ranges = dict(model_params["influent_state_ranges"])
    operational_ranges = dict(model_params["operational_ranges"])

    _validate_unique_names(measured_output_columns, "measured_output_columns")
    _validate_unique_names(process_names, "process names")
    _validate_unique_names(operational_columns, "operational_columns")

    if len(process_types) != len(process_names):
        raise ValueError("asm2d_tsn_simulation process_types must align with the configured process list.")

    missing_state_ranges = [state_name for state_name in state_columns if state_name not in influent_state_ranges]
    if missing_state_ranges:
        missing_display = ", ".join(missing_state_ranges)
        raise KeyError(f"asm2d_tsn_simulation missing influent_state_ranges for: {missing_display}")

    missing_operational_ranges = [name for name in operational_columns if name not in operational_ranges]
    if missing_operational_ranges:
        missing_display = ", ".join(missing_operational_ranges)
        raise KeyError(f"asm2d_tsn_simulation missing operational_ranges for: {missing_display}")

    return {
        "workbook_config": workbook_config,
        "solver": solver,
        "state_columns": state_columns,
        "measured_output_columns": measured_output_columns,
        "process_names": process_names,
        "process_types": process_types,
        "operational_columns": operational_columns,
        "influent_state_ranges": influent_state_ranges,
        "operational_ranges": operational_ranges,
    }


def _validate_solver_config(model_params: Mapping[str, Any]) -> dict[str, Any]:
    if "solver" not in model_params:
        raise KeyError("asm2d_tsn_simulation must define a solver section.")

    solver = dict(model_params["solver"])
    required_keys = (
        "lower_bound",
        "upper_bound",
        "initial_guess_floor",
        "warm_start_previous_weight",
        "warm_start_influent_weight",
        "initial_s_a_fraction",
        "initial_s_f_fraction",
        "initial_heterotroph_to_xs_ratio",
        "initial_pao_to_pp_ratio",
        "initial_aob_to_nh4_ratio",
        "initial_nob_to_no2_ratio",
        "multistart_s_a_fraction",
        "multistart_s_f_fraction",
        "multistart_heterotroph_to_xs_ratio",
        "multistart_pao_to_pp_ratio",
        "multistart_aob_to_nh4_ratio",
        "multistart_nob_to_no2_ratio",
        "dynamic_relaxation_days",
        "dynamic_absolute_tolerance",
        "dynamic_relative_tolerance",
        "dynamic_max_step",
        "residual_tolerance",
        "variable_tolerance",
        "gradient_tolerance",
        "acceptance_residual_max",
        "max_nfev",
    )

    for key in required_keys:
        if key not in solver:
            raise KeyError(f"asm2d_tsn_simulation solver missing '{key}'.")
        float(solver[key])

    lower_bound = float(solver["lower_bound"])
    upper_bound = float(solver["upper_bound"])
    initial_guess_floor = float(solver["initial_guess_floor"])
    if lower_bound < 0.0:
        raise ValueError("asm2d_tsn_simulation solver lower_bound must be non-negative.")
    if upper_bound <= lower_bound:
        raise ValueError("asm2d_tsn_simulation solver upper_bound must exceed lower_bound.")
    if not (lower_bound <= initial_guess_floor <= upper_bound):
        raise ValueError(
            "asm2d_tsn_simulation solver initial_guess_floor must lie between lower_bound and upper_bound."
        )

    previous_weight = float(solver["warm_start_previous_weight"])
    influent_weight = float(solver["warm_start_influent_weight"])
    if previous_weight < 0.0 or influent_weight < 0.0:
        raise ValueError("asm2d_tsn_simulation warm-start weights must be non-negative.")
    if previous_weight + influent_weight <= 0.0:
        raise ValueError("asm2d_tsn_simulation warm-start weights must sum to a positive value.")

    return solver


def _validate_unique_names(names: list[str], name_type: str) -> None:
    if not names:
        raise ValueError(f"{name_type} must not be empty.")

    duplicates = sorted({name for name in names if names.count(name) > 1})
    if duplicates:
        duplicate_display = ", ".join(duplicates)
        raise ValueError(f"asm2d_tsn_simulation {name_type} contains duplicates: {duplicate_display}")


def _build_parameter_reference_map(parameter_rows: list[Mapping[str, Any]]) -> dict[str, str]:
    value_column_letter = get_column_letter(PARAMETER_VALUE_COLUMN_INDEX)
    parameter_refs: dict[str, str] = {}

    for row_number, parameter_row in enumerate(parameter_rows, start=2):
        excel_name = str(parameter_row["excel_name"])
        parameter_refs[excel_name] = f"'{PARAMETER_SHEET_NAME}'!${value_column_letter}${row_number}"

    return parameter_refs


def _build_parameter_value_map(parameter_rows: list[Mapping[str, Any]]) -> dict[str, float]:
    return {str(parameter_row["excel_name"]): float(parameter_row["value"]) for parameter_row in parameter_rows}


def _evaluate_numeric_expression(expression: str | float | int, parameter_values: Mapping[str, float]) -> float:
    if isinstance(expression, (int, float)):
        return float(expression)

    formatted_expression = str(expression).format_map(parameter_values)
    return float(eval(formatted_expression, {"__builtins__": {}}, {}))


def _build_state_index(state_columns: list[str]) -> dict[str, int]:
    return {name: position for position, name in enumerate(state_columns)}


def _generate_lhs_candidate_pool(
    seed: int,
    n_points: int,
    ordered_names: list[str],
    ranges: Mapping[str, Any],
) -> np.ndarray:
    """Generate n_points Latin Hypercube samples scaled to the configured parameter ranges.

    For degenerate ranges where lower == upper the returned column is the constant
    lower-bound value.  An empty array is returned when n_points or the number of
    dimensions is zero.
    """
    from scipy.stats.qmc import LatinHypercube, scale as qmc_scale

    n_dims = len(ordered_names)
    if n_points == 0 or n_dims == 0:
        return np.zeros((n_points, n_dims), dtype=float)

    lower_bounds = np.array([float(ranges[name][0]) for name in ordered_names], dtype=float)
    upper_bounds = np.array([float(ranges[name][1]) for name in ordered_names], dtype=float)

    degenerate_mask = lower_bounds == upper_bounds
    active_indices = np.where(~degenerate_mask)[0]

    result = np.empty((n_points, n_dims), dtype=float)
    # Fill constant values for degenerate dimensions
    for idx in np.where(degenerate_mask)[0]:
        result[:, idx] = lower_bounds[idx]

    if len(active_indices) == 0:
        return result

    sampler = LatinHypercube(d=len(active_indices), seed=seed)
    unit_samples = sampler.random(n=n_points)
    scaled = qmc_scale(unit_samples, lower_bounds[active_indices], upper_bounds[active_indices])
    result[:, active_indices] = scaled
    return result


def _monod(numerator: float, half_saturation: float) -> float:
    return float(numerator) / max(float(numerator) + float(half_saturation), 1e-9)


def _ratio(numerator: float, denominator: float) -> float:
    return float(numerator) / max(float(denominator), 1e-9)


def _share(numerator: float, denominator_a: float, denominator_b: float) -> float:
    return float(numerator) / max(float(denominator_a) + float(denominator_b), 1e-9)


def _build_influent_state_sample(
    sampled_state: np.ndarray,
) -> np.ndarray:
    return np.maximum(sampled_state.copy(), 0.0)


def _build_initial_guess(
    influent_state: np.ndarray,
    hrt_hours: float,
    aeration: float,
    state_index: Mapping[str, int],
    model_params: Mapping[str, Any],
    previous_solution: np.ndarray | None = None,
) -> np.ndarray:
    solver = model_params["solver"]
    aeration_model = model_params["aeration_model"]
    lower_floor = float(solver["initial_guess_floor"])
    upper_bound = float(solver["upper_bound"])
    guess = np.clip(influent_state.copy(), lower_floor, upper_bound)

    if previous_solution is not None:
        previous_weight = float(solver["warm_start_previous_weight"])
        influent_weight = float(solver["warm_start_influent_weight"])
        weight_total = max(previous_weight + influent_weight, 1e-9)
        guess = np.clip(
            (previous_weight * previous_solution + influent_weight * guess) / weight_total,
            lower_floor,
            upper_bound,
        )

    guess[state_index["S_A"]] *= float(solver["initial_s_a_fraction"])
    guess[state_index["S_F"]] *= float(solver["initial_s_f_fraction"])
    guess[state_index["X_H"]] = max(
        guess[state_index["X_H"]],
        guess[state_index["X_S"]] * float(solver["initial_heterotroph_to_xs_ratio"]),
    )
    guess[state_index["X_PAO"]] = max(
        guess[state_index["X_PAO"]],
        guess[state_index["X_PP"]] * float(solver["initial_pao_to_pp_ratio"]),
    )
    guess[state_index["X_AOB"]] = max(
        guess[state_index["X_AOB"]],
        guess[state_index["S_NH4"]] * float(solver["initial_aob_to_nh4_ratio"]),
    )
    guess[state_index["X_NOB"]] = max(
        guess[state_index["X_NOB"]],
        guess[state_index["S_NO2"]] * float(solver["initial_nob_to_no2_ratio"]),
    )

    dilution_rate = 24.0 / max(float(hrt_hours), 1e-6)
    kla = float(aeration_model["kla_base"]) + float(aeration_model["kla_per_aeration"]) * max(float(aeration), 0.0)
    do_saturation = float(aeration_model["do_saturation"])
    guess[state_index["S_O"]] = np.clip(
        (dilution_rate * guess[state_index["S_O"]] + kla * do_saturation) / max(dilution_rate + kla, 1e-9),
        lower_floor,
        do_saturation,
    )

    return guess


def _build_multistart_guess(
    initial_guess: np.ndarray,
    influent_state: np.ndarray,
    state_index: Mapping[str, int],
    model_params: Mapping[str, Any],
) -> np.ndarray:
    solver = model_params["solver"]
    multistart_guess = initial_guess.copy()
    multistart_guess[state_index["S_A"]] *= float(solver["multistart_s_a_fraction"])
    multistart_guess[state_index["S_F"]] *= float(solver["multistart_s_f_fraction"])
    multistart_guess[state_index["X_H"]] = max(
        multistart_guess[state_index["X_H"]],
        influent_state[state_index["X_S"]] * float(solver["multistart_heterotroph_to_xs_ratio"]),
    )
    multistart_guess[state_index["X_PAO"]] = max(
        multistart_guess[state_index["X_PAO"]],
        influent_state[state_index["X_PP"]] * float(solver["multistart_pao_to_pp_ratio"]),
    )
    multistart_guess[state_index["X_AOB"]] = max(
        multistart_guess[state_index["X_AOB"]],
        influent_state[state_index["S_NH4"]] * float(solver["multistart_aob_to_nh4_ratio"]),
    )
    multistart_guess[state_index["X_NOB"]] = max(
        multistart_guess[state_index["X_NOB"]],
        influent_state[state_index["S_NO2"]] * float(solver["multistart_nob_to_no2_ratio"]),
    )
    return np.clip(
        multistart_guess,
        float(solver["lower_bound"]),
        float(solver["upper_bound"]),
    )


def _compute_process_rates(
    state: np.ndarray,
    model_params: Mapping[str, Any],
    state_index: Mapping[str, int],
    parameter_values: Mapping[str, float],
) -> np.ndarray:
    s_a = state[state_index["S_A"]]
    s_f = state[state_index["S_F"]]
    s_nh4 = state[state_index["S_NH4"]]
    s_no2 = state[state_index["S_NO2"]]
    s_no3 = state[state_index["S_NO3"]]
    s_po4 = state[state_index["S_PO4"]]
    s_nox = s_no2 + s_no3
    s_alk = state[state_index["S_ALK"]]
    s_o = state[state_index["S_O"]]
    x_s = state[state_index["X_S"]]
    x_h = state[state_index["X_H"]]
    x_pao = state[state_index["X_PAO"]]
    x_pp = state[state_index["X_PP"]]
    x_pha = state[state_index["X_PHA"]]
    x_aob = state[state_index["X_AOB"]]
    x_nob = state[state_index["X_NOB"]]
    x_meoh = state[state_index["X_MeOH"]]
    x_mep = state[state_index["X_MeP"]]

    xs_to_xh = _ratio(x_s, x_h)
    hydrolysis_availability = _monod(xs_to_xh, parameter_values["K_X"])
    nitrate_share = _share(s_no3, s_no3, s_no2)
    nitrite_share = _share(s_no2, s_no3, s_no2)

    oxygen_hyd = _monod(s_o, parameter_values["K_O_hyd"])
    oxygen_hyd_limitation = parameter_values["K_O_hyd"] / max(parameter_values["K_O_hyd"] + s_o, 1e-9)
    oxygen_h = _monod(s_o, parameter_values["K_O_H"])
    oxygen_h_limitation = parameter_values["K_O_H"] / max(parameter_values["K_O_H"] + s_o, 1e-9)
    oxygen_pao = _monod(s_o, parameter_values["K_O_PAO"])
    oxygen_pao_limitation = parameter_values["K_O_PAO"] / max(parameter_values["K_O_PAO"] + s_o, 1e-9)

    alk_h = _monod(s_alk, parameter_values["K_ALK_H"])
    alk_pao = _monod(s_alk, parameter_values["K_ALK_PAO"])
    alk_nit = _monod(s_alk, parameter_values["K_ALK_nit"])
    alk_chem = _monod(s_alk, parameter_values["K_ALK_chem"])

    ammonium_h = _monod(s_nh4, parameter_values["K_NH4_H"])
    phosphate_h = _monod(s_po4, parameter_values["K_PO4_H"])
    ammonium_pao = _monod(s_nh4, parameter_values["K_NH4_PAO"])
    phosphate_pao = _monod(s_po4, parameter_values["K_PO4_PAO"])
    phosphate_storage = _monod(s_po4, parameter_values["K_PS"])
    phosphate_nit = _monod(s_po4, parameter_values["K_PO4_nit"])

    substrate_f = _monod(s_f, parameter_values["K_F"])
    substrate_fe = _monod(s_f, parameter_values["K_fe"])
    substrate_a = _monod(s_a, parameter_values["K_A"])

    pao_ratio_pp = _ratio(x_pp, x_pao)
    pao_ratio_pha = _ratio(x_pha, x_pao)
    pp_capacity = max(parameter_values["K_MAX"] - pao_ratio_pp, 0.0)
    r_pp = _monod(pao_ratio_pp, parameter_values["K_PP"])
    r_pha = _monod(pao_ratio_pha, parameter_values["K_PHA"])
    c_pp = pp_capacity / max(parameter_values["K_IPP"] + pp_capacity, 1e-9)

    process_rates = np.array(
        [
            parameter_values["K_H"] * oxygen_hyd * hydrolysis_availability * x_h,
            parameter_values["eta_hyd_NO2"]
            * parameter_values["K_H"]
            * oxygen_hyd_limitation
            * _monod(s_no2, parameter_values["K_NO2_hyd"])
            * nitrite_share
            * hydrolysis_availability
            * x_h,
            parameter_values["eta_hyd_NO3"]
            * parameter_values["K_H"]
            * oxygen_hyd_limitation
            * _monod(s_no3, parameter_values["K_NO3_hyd"])
            * nitrate_share
            * hydrolysis_availability
            * x_h,
            parameter_values["eta_hyd_fe"]
            * parameter_values["K_H"]
            * oxygen_hyd_limitation
            * (parameter_values["K_NOX_hyd"] / max(parameter_values["K_NOX_hyd"] + s_nox, 1e-9))
            * hydrolysis_availability
            * x_h,
            parameter_values["mu_H"]
            * oxygen_h
            * substrate_f
            * _share(s_f, s_a, s_f)
            * ammonium_h
            * phosphate_h
            * alk_h
            * x_h,
            parameter_values["mu_H"]
            * oxygen_h
            * substrate_a
            * _share(s_a, s_a, s_f)
            * ammonium_h
            * phosphate_h
            * alk_h
            * x_h,
            parameter_values["mu_H"]
            * oxygen_h_limitation
            * substrate_f
            * _share(s_f, s_a, s_f)
            * ammonium_h
            * phosphate_h
            * alk_h
            * parameter_values["eta_H_NO3"]
            * _monod(s_no3, parameter_values["K_NO3_H"])
            * nitrate_share
            * x_h,
            parameter_values["mu_H"]
            * oxygen_h_limitation
            * substrate_f
            * _share(s_f, s_a, s_f)
            * ammonium_h
            * phosphate_h
            * alk_h
            * parameter_values["eta_H_NO2"]
            * _monod(s_no2, parameter_values["K_NO2_H"])
            * nitrite_share
            * x_h,
            parameter_values["mu_H"]
            * oxygen_h_limitation
            * substrate_a
            * _share(s_a, s_a, s_f)
            * ammonium_h
            * phosphate_h
            * alk_h
            * parameter_values["eta_H_NO3"]
            * _monod(s_no3, parameter_values["K_NO3_H"])
            * nitrate_share
            * x_h,
            parameter_values["mu_H"]
            * oxygen_h_limitation
            * substrate_a
            * _share(s_a, s_a, s_f)
            * ammonium_h
            * phosphate_h
            * alk_h
            * parameter_values["eta_H_NO2"]
            * _monod(s_no2, parameter_values["K_NO2_H"])
            * nitrite_share
            * x_h,
            parameter_values["q_Fe"]
            * parameter_values["mu_H"]
            * oxygen_h_limitation
            * (parameter_values["K_NOX_H"] / max(parameter_values["K_NOX_H"] + s_nox, 1e-9))
            * substrate_fe
            * alk_h
            * x_h,
            parameter_values["b_H"] * x_h,
            parameter_values["q_PHA"] * substrate_a * alk_pao * r_pp * x_pao,
            parameter_values["q_PP"] * oxygen_pao * phosphate_storage * alk_pao * r_pha * c_pp * x_pao,
            parameter_values["q_PP"]
            * oxygen_pao_limitation
            * phosphate_storage
            * alk_pao
            * r_pha
            * c_pp
            * parameter_values["eta_PAO_NO3"]
            * _monod(s_no3, parameter_values["K_NO3_PAO"])
            * nitrate_share
            * x_pao,
            parameter_values["q_PP"]
            * oxygen_pao_limitation
            * phosphate_storage
            * alk_pao
            * r_pha
            * c_pp
            * parameter_values["eta_PAO_NO2"]
            * _monod(s_no2, parameter_values["K_NO2_PAO"])
            * nitrite_share
            * x_pao,
            parameter_values["mu_PAO"] * oxygen_pao * ammonium_pao * phosphate_pao * r_pha * alk_pao * x_pao,
            parameter_values["mu_PAO"]
            * oxygen_pao_limitation
            * ammonium_pao
            * phosphate_pao
            * r_pha
            * alk_pao
            * parameter_values["eta_PAO_NO3"]
            * _monod(s_no3, parameter_values["K_NO3_PAO"])
            * nitrate_share
            * x_pao,
            parameter_values["mu_PAO"]
            * oxygen_pao_limitation
            * ammonium_pao
            * phosphate_pao
            * r_pha
            * alk_pao
            * parameter_values["eta_PAO_NO2"]
            * _monod(s_no2, parameter_values["K_NO2_PAO"])
            * nitrite_share
            * x_pao,
            parameter_values["b_PAO"] * x_pao * alk_pao,
            parameter_values["b_PP"] * x_pp * alk_pao,
            parameter_values["b_PHA"] * x_pha * alk_pao,
            parameter_values["mu_AOB"]
            * _monod(s_o, parameter_values["K_O_AOB"])
            * _monod(s_nh4, parameter_values["K_NH4_AOB"])
            * phosphate_nit
            * alk_nit
            * x_aob,
            parameter_values["mu_NOB"]
            * _monod(s_o, parameter_values["K_O_NOB"])
            * _monod(s_no2, parameter_values["K_NO2_NOB"])
            * phosphate_nit
            * alk_nit
            * x_nob,
            parameter_values["b_AOB"] * x_aob,
            parameter_values["b_NOB"] * x_nob,
            parameter_values["k_PRE"] * s_po4 * x_meoh,
            parameter_values["k_RED"] * x_mep * alk_chem,
        ],
        dtype=float,
    )
    return np.clip(process_rates, 0.0, float(model_params["hyperparameters"]["max_process_rate"]))


def _compute_aeration_flux(state: np.ndarray, aeration: float, state_index: Mapping[str, int], model_params: Mapping[str, Any]) -> float:
    aeration_model = model_params["aeration_model"]
    kla = float(aeration_model["kla_base"]) + float(aeration_model["kla_per_aeration"]) * max(float(aeration), 0.0)
    do_saturation = float(aeration_model["do_saturation"])
    return kla * (do_saturation - state[state_index["S_O"]])


def _steady_state_residuals(
    state: np.ndarray,
    influent_state: np.ndarray,
    hrt_hours: float,
    aeration: float,
    matrix_bundle: Mapping[str, Any],
    model_params: Mapping[str, Any],
    parameter_values: Mapping[str, float],
) -> np.ndarray:
    dilution_rate = 24.0 / max(float(hrt_hours), 1e-6)
    state_index = dict(matrix_bundle["state_index"])
    residual = dilution_rate * (influent_state - state)
    process_rates = _compute_process_rates(state, model_params, state_index, parameter_values)
    residual += process_rates @ np.asarray(matrix_bundle["petersen_matrix"], dtype=float)
    residual[state_index["S_O"]] += _compute_aeration_flux(state, aeration, state_index, model_params)
    return residual


def simulate_asm2d_tsn_steady_state(
    *,
    influent_state: np.ndarray,
    hrt_hours: float,
    aeration: float,
    model_params: Mapping[str, Any],
    matrix_bundle: Mapping[str, Any] | None = None,
    previous_solution: np.ndarray | None = None,
    enforce_acceptance: bool = True,
) -> tuple[np.ndarray, dict[str, float | bool | int]]:
    """Solve a single mechanistic steady-state ASM2D-TSN operating point."""

    matrix_bundle = matrix_bundle if matrix_bundle is not None else get_asm2d_tsn_matrices(model_params)
    resolved_measured_output_columns = (
        list(matrix_bundle["measured_output_columns"])
        if "measured_output_columns" in matrix_bundle
        else None
    )
    runtime = _validate_runtime_structure(
        model_params,
        measured_output_columns=resolved_measured_output_columns,
    )
    parameter_values = _build_parameter_value_map(runtime["workbook_config"]["parameters"])
    state_columns = list(runtime["state_columns"])
    state_index = dict(matrix_bundle["state_index"])
    solver = runtime["solver"]
    lower_bounds = np.full(len(state_columns), float(solver["lower_bound"]), dtype=float)
    upper_bounds = np.full(len(state_columns), float(solver["upper_bound"]), dtype=float)
    initial_guess = _build_initial_guess(
        influent_state,
        hrt_hours,
        aeration,
        state_index,
        model_params,
        previous_solution=previous_solution,
    )
    candidate_guesses = [initial_guess, _build_multistart_guess(initial_guess, influent_state, state_index, model_params)]
    candidate_labels = ["initial", "multistart"]

    best_result = None
    best_residual_max = np.inf
    best_result_label = "initial"
    candidate_residuals: dict[str, float] = {}
    for candidate_label, candidate_guess in zip(candidate_labels, candidate_guesses, strict=True):
        result = least_squares(
            _steady_state_residuals,
            candidate_guess,
            bounds=(lower_bounds, upper_bounds),
            xtol=float(solver["variable_tolerance"]),
            ftol=float(solver["residual_tolerance"]),
            gtol=float(solver["gradient_tolerance"]),
            max_nfev=int(solver["max_nfev"]),
            args=(influent_state, hrt_hours, aeration, matrix_bundle, model_params, parameter_values),
        )
        residual_max = float(np.max(np.abs(result.fun)))
        candidate_residuals[candidate_label] = residual_max
        if residual_max < best_residual_max:
            best_result = result
            best_residual_max = residual_max
            best_result_label = candidate_label

    dynamic_relaxation_used = best_residual_max > float(solver["acceptance_residual_max"])
    dynamic_relaxation_improved = False
    if best_residual_max > float(solver["acceptance_residual_max"]):
        dynamic_result = solve_ivp(
            lambda _time, values: _steady_state_residuals(
                values,
                influent_state,
                hrt_hours,
                aeration,
                matrix_bundle,
                model_params,
                parameter_values,
            ),
            (0.0, float(solver["dynamic_relaxation_days"])),
            np.clip(candidate_guesses[-1], lower_bounds, upper_bounds),
            method="BDF",
            atol=float(solver["dynamic_absolute_tolerance"]),
            rtol=float(solver["dynamic_relative_tolerance"]),
            max_step=float(solver["dynamic_max_step"]),
        )
        if dynamic_result.success:
            relaxed_guess = np.clip(dynamic_result.y[:, -1], lower_bounds, upper_bounds)
            result = least_squares(
                _steady_state_residuals,
                relaxed_guess,
                bounds=(lower_bounds, upper_bounds),
                xtol=float(solver["variable_tolerance"]),
                ftol=float(solver["residual_tolerance"]),
                gtol=float(solver["gradient_tolerance"]),
                max_nfev=int(solver["max_nfev"]),
                args=(influent_state, hrt_hours, aeration, matrix_bundle, model_params, parameter_values),
            )
            residual_max = float(np.max(np.abs(result.fun)))
            if residual_max < best_residual_max:
                best_result = result
                best_residual_max = residual_max
                best_result_label = "dynamic_relaxation"
                dynamic_relaxation_improved = True

    assert best_result is not None
    result = best_result
    accepted = bool(result.success and best_residual_max <= float(solver["acceptance_residual_max"]))
    diagnostics: dict[str, float | bool | int] = {
        "success": bool(result.success),
        "accepted": accepted,
        "status": int(result.status),
        "nfev": int(result.nfev),
        "residual_l2": float(np.linalg.norm(result.fun)),
        "residual_max": best_residual_max,
        "acceptance_threshold": float(solver["acceptance_residual_max"]),
        "initial_residual_max": float(candidate_residuals["initial"]),
        "multistart_residual_max": float(candidate_residuals["multistart"]),
        "selected_strategy": best_result_label,
        "dynamic_relaxation_used": dynamic_relaxation_used,
        "dynamic_relaxation_improved": dynamic_relaxation_improved,
    }

    if enforce_acceptance and (not accepted):
        raise RuntimeError(
            "asm2d_tsn_simulation steady-state solve failed: "
            f"success={result.success}, status={result.status}, residual_max={diagnostics['residual_max']:.3e}"
        )

    return result.x, diagnostics


def _compute_measured_output_values(state: np.ndarray, matrix_bundle: Mapping[str, Any]) -> np.ndarray:
    return np.asarray(matrix_bundle["composition_matrix"], dtype=float) @ state


def _generate_asm2d_tsn_dataset_chunk(
    *,
    chunk_start: int,
    chunk_size: int,
    chunk_seed: int,
    model_params: Mapping[str, Any],
    matrix_bundle: Mapping[str, Any],
    runtime: Mapping[str, Any],
    collect_debug_data: bool = False,
    progress_bar=None,
) -> tuple[int, np.ndarray, np.ndarray, np.ndarray, np.ndarray, list[dict[str, Any]]]:
    configured_hyperparameters = model_params["hyperparameters"]
    max_sample_attempts = int(configured_hyperparameters["max_sample_attempts"])
    state_columns = list(runtime["state_columns"])
    operational_columns = list(runtime["operational_columns"])
    measured_output_columns = list(runtime["measured_output_columns"])
    state_index = dict(matrix_bundle["state_index"])
    parameter_values = _build_parameter_value_map(runtime["workbook_config"]["parameters"])

    # Pre-generate a joint LHS candidate pool covering all retry slots.
    # Influent-state and operational columns are combined into a single LHS design
    # so their stratification is jointly controlled.  Sequential consumption means
    # each retry draws the next LHS point rather than an independent uniform draw.
    total_candidates = chunk_size * max_sample_attempts
    all_columns = state_columns + operational_columns
    all_ranges = {**runtime["influent_state_ranges"], **runtime["operational_ranges"]}
    candidate_pool = _generate_lhs_candidate_pool(chunk_seed, total_candidates, all_columns, all_ranges)
    n_state = len(state_columns)

    influent_states = np.zeros((chunk_size, len(state_columns)), dtype=float)
    operational = np.zeros((chunk_size, len(operational_columns)), dtype=float)
    effluent_states = np.zeros((chunk_size, len(state_columns)), dtype=float)
    measured_outputs = np.zeros((chunk_size, len(measured_output_columns)), dtype=float)
    solver_diagnostics: list[dict[str, Any]] = []

    previous_solution: np.ndarray | None = None
    for local_index in range(chunk_size):
        last_error: RuntimeError | None = None
        for attempt_index in range(max_sample_attempts):
            candidate_index = local_index * max_sample_attempts + attempt_index
            sampled_state = candidate_pool[candidate_index, :n_state]
            candidate_influent = _build_influent_state_sample(sampled_state)
            candidate_operational = candidate_pool[candidate_index, n_state:]

            try:
                effluent_state, diagnostics = simulate_asm2d_tsn_steady_state(
                    influent_state=candidate_influent,
                    hrt_hours=float(candidate_operational[0]),
                    aeration=float(candidate_operational[1]),
                    model_params=model_params,
                    matrix_bundle=matrix_bundle,
                    previous_solution=previous_solution,
                )
            except RuntimeError as error:
                last_error = error
                continue

            if not diagnostics["accepted"]:
                last_error = RuntimeError(
                    "asm2d_tsn_simulation steady-state solve did not satisfy the configured acceptance threshold."
                )
                continue

            if not np.all(np.isfinite(effluent_state)):
                last_error = RuntimeError("asm2d_tsn_simulation produced non-finite effluent states.")
                continue

            previous_solution = effluent_state
            influent_states[local_index] = candidate_influent
            operational[local_index] = candidate_operational
            effluent_states[local_index] = effluent_state
            measured_outputs[local_index] = _compute_measured_output_values(effluent_state, matrix_bundle)
            if collect_debug_data:
                diagnostic_record = dict(diagnostics)
                diagnostic_record["sample_index"] = chunk_start + local_index
                diagnostic_record["attempt_count"] = attempt_index + 1
                diagnostic_record["HRT"] = float(candidate_operational[0])
                diagnostic_record["Aeration"] = float(candidate_operational[1])
                solver_diagnostics.append(diagnostic_record)
            if progress_bar is not None:
                progress_bar.update(1)
            break
        else:
            raise RuntimeError(
                "asm2d_tsn_simulation failed to generate a valid sample after "
                f"{max_sample_attempts} attempts at sample index {chunk_start + local_index}."
            ) from last_error

    return chunk_start, influent_states, operational, effluent_states, measured_outputs, solver_diagnostics


def _summarize_asm2d_tsn_solver_diagnostics(
    solver_diagnostics: pd.DataFrame,
    acceptance_threshold: float,
) -> dict[str, Any]:
    if solver_diagnostics.empty:
        return {
            "sample_count": 0,
            "accepted_count": 0,
            "accepted_rate": 0.0,
            "acceptance_threshold": acceptance_threshold,
            "multistart_selected_rate": 0.0,
            "dynamic_relaxation_used_rate": 0.0,
            "dynamic_relaxation_improved_rate": 0.0,
            "residual_max_quantiles": {},
            "nfev_quantiles": {},
            "selected_strategy_counts": {},
        }

    residual_max_quantiles = {
        "q50": float(solver_diagnostics["residual_max"].quantile(0.50)),
        "q90": float(solver_diagnostics["residual_max"].quantile(0.90)),
        "q95": float(solver_diagnostics["residual_max"].quantile(0.95)),
        "q99": float(solver_diagnostics["residual_max"].quantile(0.99)),
        "max": float(solver_diagnostics["residual_max"].max()),
    }
    nfev_quantiles = {
        "q50": float(solver_diagnostics["nfev"].quantile(0.50)),
        "q90": float(solver_diagnostics["nfev"].quantile(0.90)),
        "q95": float(solver_diagnostics["nfev"].quantile(0.95)),
        "max": float(solver_diagnostics["nfev"].max()),
    }
    return {
        "sample_count": int(len(solver_diagnostics)),
        "accepted_count": int(solver_diagnostics["accepted"].sum()),
        "accepted_rate": float(solver_diagnostics["accepted"].mean()),
        "acceptance_threshold": float(acceptance_threshold),
        "multistart_selected_rate": float((solver_diagnostics["selected_strategy"] == "multistart").mean()),
        "dynamic_relaxation_used_rate": float(solver_diagnostics["dynamic_relaxation_used"].mean()),
        "dynamic_relaxation_improved_rate": float(solver_diagnostics["dynamic_relaxation_improved"].mean()),
        "mean_attempt_count": float(solver_diagnostics.get("attempt_count", pd.Series([1])).mean()),
        "max_attempt_count": int(solver_diagnostics.get("attempt_count", pd.Series([1])).max()),
        "residual_max_quantiles": residual_max_quantiles,
        "nfev_quantiles": nfev_quantiles,
        "selected_strategy_counts": {
            str(name): int(count)
            for name, count in solver_diagnostics["selected_strategy"].value_counts().items()
        },
    }


def _resolve_parallel_workers(requested_workers: int, sample_count: int) -> int:
    if requested_workers < 0:
        raise ValueError("parallel_workers must be greater than or equal to 0.")

    if sample_count <= 1:
        return 1

    available_workers = os.cpu_count() or 1
    if requested_workers == 0:
        requested_workers = max(available_workers - 1, 1)

    return min(max(requested_workers, 1), available_workers, sample_count)


def _resolve_parallel_chunk_size(requested_chunk_size: int) -> int:
    if requested_chunk_size < 1:
        raise ValueError("parallel_chunk_size must be at least 1.")

    return requested_chunk_size


def _write_parameter_sheet(worksheet, parameter_rows: list[Mapping[str, Any]]) -> None:
    worksheet.freeze_panes = "A2"
    headers = ["category", "symbol", "excel_name", "description", "value", "unit"]
    _write_header_row(worksheet, headers)

    previous_category = None
    for row_number, parameter_row in enumerate(parameter_rows, start=2):
        current_category = str(parameter_row["category"])
        row_values = [
            current_category,
            str(parameter_row["symbol"]),
            str(parameter_row["excel_name"]),
            str(parameter_row["description"]),
            float(parameter_row["value"]),
            str(parameter_row["unit"]),
        ]

        for column_number, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_number, column=column_number, value=value)
            if column_number == PARAMETER_VALUE_COLUMN_INDEX:
                cell.number_format = "0.###############"

        if current_category != previous_category:
            for column_number in range(1, len(headers) + 1):
                worksheet.cell(row=row_number, column=column_number).fill = SECTION_FILL
        previous_category = current_category


def _write_stoichiometric_sheet(
    worksheet,
    workbook_config: Mapping[str, Any],
    parameter_refs: Mapping[str, str],
) -> None:
    worksheet.freeze_panes = "C2"
    state_columns = list(workbook_config["state_columns"])
    processes = list(workbook_config["processes"])
    headers = ["process_index", "process"] + state_columns
    _write_header_row(worksheet, headers)
    state_column_index = {state_name: position for position, state_name in enumerate(state_columns, start=3)}

    for row_number, process in enumerate(processes, start=2):
        worksheet.cell(row=row_number, column=1, value=int(process["index"]))
        worksheet.cell(row=row_number, column=2, value=str(process["name"]))
        direct_coefficients = STOICHIOMETRIC_COEFFICIENTS[row_number - 2]["coefficients"]

        for state_name in state_columns:
            column_number = state_column_index[state_name]
            cell = worksheet.cell(row=row_number, column=column_number)

            if state_name in direct_coefficients:
                cell.value = _format_formula(direct_coefficients[state_name], parameter_refs)
                continue

            if state_name == "S_NH4":
                cell.value = _build_weighted_formula(
                    row_number,
                    state_column_index,
                    NITROGEN_CONTINUITY_TERMS,
                    parameter_refs,
                    negate=True,
                )
                continue

            if state_name == "S_PO4":
                cell.value = _build_weighted_formula(
                    row_number,
                    state_column_index,
                    PHOSPHORUS_CONTINUITY_TERMS,
                    parameter_refs,
                    negate=True,
                )
                continue

            if state_name == "S_ALK":
                cell.value = _build_alkalinity_formula(row_number, state_column_index)
                continue

def _write_composition_sheet(
    worksheet,
    workbook_config: Mapping[str, Any],
    parameter_refs: Mapping[str, str],
) -> None:
    worksheet.freeze_panes = "D2"
    dissolved_state_columns = list(workbook_config["dissolved_state_columns"])
    particulate_state_columns = list(workbook_config["particulate_state_columns"])
    state_columns = list(workbook_config["state_columns"])
    legacy_composite_variables = workbook_config.get("composite_variables")
    composite_variables = _resolve_workbook_composite_variables(
        state_columns,
        legacy_composite_variables=(
            None if legacy_composite_variables is None else list(legacy_composite_variables)
        ),
    )
    state_units = dict(workbook_config["state_units"])
    headers = ["state_group", "state_variable", "unit"] + composite_variables
    _write_header_row(worksheet, headers)
    composite_column_index = {name: position for position, name in enumerate(composite_variables, start=4)}

    for row_number, state_name in enumerate(state_columns, start=2):
        state_group = "Dissolved" if state_name in dissolved_state_columns else "Particulate"
        worksheet.cell(row=row_number, column=1, value=state_group)
        worksheet.cell(row=row_number, column=2, value=state_name)
        worksheet.cell(row=row_number, column=3, value=state_units[state_name])

        for composite_name, expression in COMPOSITION_FORMULAS.get(state_name, {}).items():
            worksheet.cell(
                row=row_number,
                column=composite_column_index[composite_name],
                value=_format_formula(expression, parameter_refs),
            )


def _resolve_workbook_composite_variables(
    state_columns: list[str],
    *,
    legacy_composite_variables: list[str] | None,
) -> list[str]:
    composite_variables: list[str] = []
    for state_name in state_columns:
        for composite_name in COMPOSITION_FORMULAS.get(state_name, {}):
            normalized_name = str(composite_name)
            if normalized_name not in composite_variables:
                composite_variables.append(normalized_name)

    if not composite_variables:
        raise ValueError(
            "asm2d_tsn_simulation composition formulas do not define any composite output columns."
        )

    if legacy_composite_variables is not None:
        normalized_legacy = [str(name) for name in legacy_composite_variables]
        _validate_unique_names(normalized_legacy, "composite_variables")
        missing_formulas = sorted(name for name in normalized_legacy if name not in composite_variables)
        if missing_formulas:
            missing_display = ", ".join(missing_formulas)
            raise ValueError(
                "asm2d_tsn_simulation legacy workbook composite_variables contains outputs without "
                f"composition formulas: {missing_display}"
            )

    return composite_variables


def _write_header_row(worksheet, headers: list[str]) -> None:
    for column_number, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_number, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _format_formula(expression: str | float | int, parameter_refs: Mapping[str, str]) -> str:
    if isinstance(expression, (int, float)):
        return f"={expression:g}"

    formatted_expression = str(expression).format_map(parameter_refs)
    if formatted_expression.startswith("="):
        return formatted_expression

    return f"={formatted_expression}"


def _build_weighted_formula(
    row_number: int,
    state_column_index: Mapping[str, int],
    factor_terms: Mapping[str, str],
    parameter_refs: Mapping[str, str],
    *,
    negate: bool,
) -> str:
    terms: list[str] = []
    for state_name, factor_expression in factor_terms.items():
        cell_reference = f"{get_column_letter(state_column_index[state_name])}{row_number}"
        formatted_factor = factor_expression.format_map(parameter_refs)
        if formatted_factor == "1":
            terms.append(cell_reference)
        else:
            terms.append(f"{cell_reference}*({formatted_factor})")

    if not terms:
        return "=0"

    expression = "+".join(terms)
    if negate:
        return f"=-({expression})"

    return f"={expression}"


def _build_alkalinity_formula(row_number: int, state_column_index: Mapping[str, int]) -> str:
    ammonium_ref = f"{get_column_letter(state_column_index['S_NH4'])}{row_number}"
    nitrite_ref = f"{get_column_letter(state_column_index['S_NO2'])}{row_number}"
    nitrate_ref = f"{get_column_letter(state_column_index['S_NO3'])}{row_number}"
    phosphate_ref = f"{get_column_letter(state_column_index['S_PO4'])}{row_number}"
    return f"={ammonium_ref}/14-{nitrite_ref}/14-{nitrate_ref}/14+{phosphate_ref}/31"


def _auto_size_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


__all__ = [
    "build_asm2d_tsn_workbook",
    "build_asm2d_tsn_metadata",
    "create_asm2d_tsn_workbook",
    "generate_asm2d_tsn_dataset",
    "get_asm2d_tsn_matrices",
    "load_asm2d_tsn_workbook_composition",
    "load_asm2d_tsn_simulation_params",
    "resolve_asm2d_tsn_composition_cache_paths",
    "resolve_asm2d_tsn_simulation_artifact_paths",
    "resolve_asm2d_tsn_workbook_path",
    "run_asm2d_tsn_simulation",
    "simulate_asm2d_tsn_steady_state",
    "sweep_asm2d_tsn_operating_space",
]
